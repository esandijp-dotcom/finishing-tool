#!/usr/bin/env python3
"""
VFX Plate Exporter — single-file entry point.
"""

import sys
import os

# ─── Version & update config ────────────────────────────────────────────────
VERSION_URL     = "https://raw.githubusercontent.com/esandijp-dotcom/finishing-tool/main/version.json"
DOWNLOAD_URL    = "https://raw.githubusercontent.com/esandijp-dotcom/finishing-tool/main/main.py"

def _load_local_version():
    """Read version from version.json next to this script."""
    try:
        import json as _json
        _vpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "version.json")
        with open(_vpath) as _f:
            return _json.load(_f).get("version", "1.0")
    except Exception:
        return "1.0"

APP_VERSION = _load_local_version()
# ─────────────────────────────────────────────────────────────────────────────
import re
import subprocess
import threading
import time
import glob
import tempfile
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional
from collections import defaultdict
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

os.environ["PATH"] += os.pathsep + "/opt/homebrew/bin"
try:
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = "/opt/homebrew/bin/tesseract"
except ImportError:
    pass

# ═══════════════════════════════════════════════════════════════════════════════
# RESOLVE CONNECTOR
# ═══════════════════════════════════════════════════════════════════════════════

def get_resolve():
    paths = [
        "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules",
        os.path.expanduser("~/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules"),
    ]
    for path in paths:
        if os.path.exists(path) and path not in sys.path:
            sys.path.append(path)
    try:
        import DaVinciResolveScript as dvr
        resolve = dvr.scriptapp("Resolve")
        if resolve is None:
            raise ConnectionError("Could not connect to DaVinci Resolve. Make sure it is open.")
        return resolve
    except ImportError:
        raise ImportError("DaVinci Resolve scripting module not found.")

def get_current_project(resolve):
    pm = resolve.GetProjectManager()
    project = pm.GetCurrentProject()
    if not project:
        raise RuntimeError("No project is currently open in DaVinci Resolve.")
    return project

def get_current_timeline(project):
    timeline = project.GetCurrentTimeline()
    if not timeline:
        raise RuntimeError("No timeline is currently open.")
    return timeline

def get_timeline_info(timeline):
    return {
        "name": timeline.GetName(),
        "frame_rate": timeline.GetSetting("timelineFrameRate"),
        "start_frame": timeline.GetStartFrame(),
        "end_frame": timeline.GetEndFrame(),
    }

def get_all_clips_on_track(timeline, track_type, track_index):
    clips = timeline.GetItemListInTrack(track_type, track_index)
    return clips or []

def get_track_count(timeline, track_type="video"):
    return timeline.GetTrackCount(track_type)

def is_clip_orange(clip):
    """Returns True for Orange and Apricot clips only. Chocolate and others are skipped."""
    return clip.GetClipColor() in ("Orange", "Apricot")

def is_clip_apricot(clip):
    return clip.GetClipColor() == "Apricot"

def is_clip_chocolate(clip):
    return clip.GetClipColor() == "Chocolate"

def is_clip_enabled(clip):
    return clip.GetProperty("Enabled") != "0"

def get_clip_start_frame(clip):
    return clip.GetStart()

def get_clip_end_frame(clip):
    return clip.GetEnd()

def get_offline_reference_path(project, show_code=""):
    """
    Find the offline reference video from the media pool by matching:
    1. Contains 'REF' in the name
    2. Contains the show code if available
    3. Duration closest to the current timeline duration
    Falls back to just REF + duration if show code doesn't match.
    """
    try:
        timeline = project.GetCurrentTimeline()
        if not timeline:
            return None

        # Get timeline duration in seconds
        fps = float(timeline.GetSetting("timelineFrameRate") or 24)
        tl_duration = (timeline.GetEndFrame() - timeline.GetStartFrame()) / fps

        # Search entire media pool
        media_pool = project.GetMediaPool()
        root = media_pool.GetRootFolder()

        candidates = []

        def search_folder(folder):
            clips = folder.GetClipList()
            if clips:
                for clip in clips:
                    props = clip.GetClipProperty()
                    name = props.get("Clip Name", "").upper()
                    file_path = props.get("File Path", "")
                    duration_str = props.get("Duration", "")

                    if not file_path:
                        continue
                    if not any(ext in file_path.lower() for ext in [".mov", ".mp4", ".mxf"]):
                        continue
                    if "REF" not in name:
                        continue

                    # Parse duration (format: HH:MM:SS:FF or frames)
                    clip_duration = None
                    try:
                        if ":" in duration_str:
                            parts = duration_str.split(":")
                            h, m, s, f = int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3])
                            clip_duration = h*3600 + m*60 + s + f/fps
                        else:
                            clip_duration = float(duration_str) / fps
                    except Exception:
                        clip_duration = None

                    duration_diff = abs(clip_duration - tl_duration) if clip_duration else 9999
                    has_show_code = show_code.upper() in name if show_code else False

                    candidates.append({
                        "path": file_path,
                        "name": name,
                        "duration_diff": duration_diff,
                        "has_show_code": has_show_code,
                    })

            subfolders = folder.GetSubFolderList()
            if subfolders:
                for sub in subfolders:
                    search_folder(sub)

        search_folder(root)

        # Also search clips used directly in the current timeline
        try:
            track_count = timeline.GetTrackCount("video")
            for t in range(1, track_count + 1):
                clips = timeline.GetItemListInTrack("video", t)
                if not clips:
                    continue
                for clip in clips:
                    media = clip.GetMediaPoolItem()
                    if not media:
                        continue
                    props = media.GetClipProperty()
                    name = props.get("Clip Name", "").upper()
                    file_path = props.get("File Path", "")
                    duration_str = props.get("Duration", "")

                    if not file_path:
                        continue
                    if not any(ext in file_path.lower() for ext in [".mov", ".mp4", ".mxf"]):
                        continue
                    if "REF" not in name:
                        continue

                    # Avoid duplicates
                    if any(c["path"] == file_path for c in candidates):
                        continue

                    clip_duration = None
                    try:
                        if ":" in duration_str:
                            parts = duration_str.split(":")
                            h, m, s, f = int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3])
                            clip_duration = h*3600 + m*60 + s + f/fps
                        else:
                            clip_duration = float(duration_str) / fps
                    except Exception:
                        clip_duration = None

                    duration_diff = abs(clip_duration - tl_duration) if clip_duration else 9999
                    has_show_code = show_code.upper() in name if show_code else False

                    candidates.append({
                        "path": file_path,
                        "name": name,
                        "duration_diff": duration_diff,
                        "has_show_code": has_show_code,
                    })
        except Exception:
            pass

        if not candidates:
            return None

        # Extract YYMMDD date from filename if present
        import re as _re
        for c in candidates:
            m = _re.search('[0-9]{6}', c["name"])
            c["date"] = int(m.group(0)) if m else 0

        # Check if timeline name contains "VFX"
        tl_name = timeline.GetName().upper()
        tl_has_vfx = "VFX" in tl_name

        # Score each candidate - all criteria must match for best score
        def score(c):
            has_show = c["has_show_code"]
            has_vfx = "VFX" in c["name"] if tl_has_vfx else True
            close_duration = c["duration_diff"] < 60  # within 60 seconds
            # Higher score = better match. All three criteria ideally True.
            match_score = -(int(has_show) + int(has_vfx) + int(close_duration))
            return (match_score, -c["date"], c["duration_diff"])

        candidates.sort(key=score)
        return candidates[0]["path"]

    except Exception:
        return None

def frame_to_timecode(frame, frame_rate):
    fps = round(float(frame_rate))
    frames = int(frame)
    ff = frames % fps
    total_seconds = frames // fps
    ss = total_seconds % 60
    mm = (total_seconds // 60) % 60
    hh = total_seconds // 3600
    return f"{hh:02d}:{mm:02d}:{ss:02d}:{ff:02d}"

def frames_to_davinci_tc(frame_number, frame_rate, timeline_start_frame, start_timecode_str):
    parts = start_timecode_str.split(":")
    sh, sm, ss, sf = int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3])
    fps = round(float(frame_rate))
    start_tc_frames = ((sh * 3600 + sm * 60 + ss) * fps) + sf
    offset = frame_number - timeline_start_frame
    total = start_tc_frames + offset
    ff = total % fps
    total_secs = total // fps
    hh = total_secs // 3600
    mm = (total_secs % 3600) // 60
    ss2 = total_secs % 60
    return f"{hh:02d}:{mm:02d}:{ss2:02d}:{ff:02d}"

def timecode_str_to_frames(timecode, frame_rate):
    fps = round(float(frame_rate))
    tc = str(timecode).strip()
    if ":" in tc:
        parts = tc.split(":")
        hh = int(parts[0])
        mm = int(parts[1]) if len(parts) > 1 else 0
        ss = int(parts[2]) if len(parts) > 2 else 0
        ff = int(parts[3]) if len(parts) > 3 else 0
    else:
        hh = int(tc)
        mm, ss, ff = 59, 50, 0
    return int((hh * 3600 + mm * 60 + ss) * fps + ff)

def clear_render_queue(project):
    project.DeleteAllRenderJobs()

# ═══════════════════════════════════════════════════════════════════════════════
# EPISODE DETECTOR
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class EpisodeMarker:
    episode_code: str
    frame_number: int
    timecode: str

def check_dependencies():
    missing = []
    try:
        import cv2
    except ImportError:
        missing.append("opencv-python")
    try:
        import pytesseract
    except ImportError:
        missing.append("pytesseract")
    result = subprocess.run(["which", "tesseract"], capture_output=True)
    if result.returncode != 0:
        missing.append("tesseract (brew install tesseract)")
    return missing

def extract_episode_code(frame):
    import cv2
    import pytesseract

    # Crop to center 40% of frame - episode title cards are almost always centered
    h, w = frame.shape[:2]
    y1 = int(h * 0.30)
    y2 = int(h * 0.70)
    x1 = int(w * 0.10)
    x2 = int(w * 0.90)
    cropped = frame[y1:y2, x1:x2]

    gray = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 128, 255, cv2.THRESH_BINARY)
    ch, cw = thresh.shape
    enlarged = cv2.resize(thresh, (cw * 2, ch * 2), interpolation=cv2.INTER_CUBIC)
    config = "--psm 6"
    text = pytesseract.image_to_string(enlarged, config=config).upper().strip()
    match = re.search('EPISODE[ \t]+([0-9]+[A-Z]?)', text)
    if match:
        ep_raw = match.group(1)
        num_match = re.match('([0-9]+)([A-Z]?)', ep_raw)
        if num_match:
            num = int(num_match.group(1))
            letter = num_match.group(2)
            return f"EP{num:02d}{letter}"
    return None

# ═══════════════════════════════════════════════════════════════════════════════
# PLATE ORGANIZER
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class PlateLayer:
    clip: object
    track_index: int
    start_frame: int
    end_frame: int
    layer_number: int
    sub_index: Optional[int] = None
    filename: str = ""

@dataclass
class VFXPlate:
    anchor_clip: object
    start_frame: int
    end_frame: int
    episode_code: str = ""
    plate_number: int = 0
    filename: str = ""
    layers: list = field(default_factory=list)

def get_vfx_plates(timeline):
    """
    Collect all orange/apricot clips from all tracks (2+).
    Group overlapping clips into plates.
    The lowest-track clip in each group is the anchor.
    """
    track_count = get_track_count(timeline, "video")

    # Collect all orange clips from all tracks
    clips_by_track = {}
    for track_idx in range(2, track_count + 1):
        clips = get_all_clips_on_track(timeline, "video", track_idx)
        vfx = [c for c in clips if is_clip_orange(c) and is_clip_enabled(c)]
        if vfx:
            clips_by_track[track_idx] = sorted(vfx, key=lambda c: get_clip_start_frame(c))

    if not clips_by_track:
        return []

    # Build a flat list of all clips with their track
    all_clips = []
    for track_idx in sorted(clips_by_track.keys()):
        for clip in clips_by_track[track_idx]:
            all_clips.append((track_idx, clip))
    all_clips.sort(key=lambda x: (get_clip_start_frame(x[1]), x[0]))

    # Group clips that overlap in time into plates
    # Anchor = lowest-track clip; all others are layers above it
    plates = []
    used = set()

    for i, (anchor_track, anchor_clip) in enumerate(all_clips):
        if id(anchor_clip) in used:
            continue
        used.add(id(anchor_clip))

        a_start = get_clip_start_frame(anchor_clip)
        a_end = get_clip_end_frame(anchor_clip)
        plate = VFXPlate(anchor_clip=anchor_clip,
                         start_frame=a_start, end_frame=a_end)

        # Find all clips on higher tracks that overlap with this anchor
        overlapping = []
        for track_idx, clip in all_clips:
            if id(clip) in used:
                continue
            if track_idx <= anchor_track:
                continue
            cs = get_clip_start_frame(clip)
            ce = get_clip_end_frame(clip)
            if cs >= a_start and ce <= a_end:
                overlapping.append((track_idx, clip))
                used.add(id(clip))

        if overlapping:
            # Count total orange (non-apricot) clips in this stack
            all_in_stack = [(anchor_track, anchor_clip)] + overlapping
            total_orange = sum(1 for _, c in all_in_stack if not is_clip_apricot(c))
            use_v_numbers = total_orange >= 2
            orange_counter = 0

            # Add anchor as first layer
            if is_clip_apricot(anchor_clip):
                plate.layers.append(PlateLayer(
                    clip=anchor_clip, track_index=anchor_track,
                    start_frame=a_start, end_frame=a_end,
                    layer_number=0, sub_index=None
                ))
            else:
                orange_counter += 1
                plate.layers.append(PlateLayer(
                    clip=anchor_clip, track_index=anchor_track,
                    start_frame=a_start, end_frame=a_end,
                    layer_number=orange_counter if use_v_numbers else -1,
                    sub_index=None
                ))

            # Add overlapping clips as layers, grouped by track
            track_clips = defaultdict(list)
            for track_idx, c in overlapping:
                track_clips[track_idx].append(c)

            for track_idx in sorted(track_clips.keys()):
                clips_on_track = track_clips[track_idx]
                for clip in clips_on_track:
                    if is_clip_apricot(clip):
                        plate.layers.append(PlateLayer(
                            clip=clip, track_index=track_idx,
                            start_frame=get_clip_start_frame(clip),
                            end_frame=get_clip_end_frame(clip),
                            layer_number=0, sub_index=None
                        ))
                    else:
                        orange_counter += 1
                        plate.layers.append(PlateLayer(
                            clip=clip, track_index=track_idx,
                            start_frame=get_clip_start_frame(clip),
                            end_frame=get_clip_end_frame(clip),
                            layer_number=orange_counter if use_v_numbers else -1,
                            sub_index=clips_on_track.index(clip) + 1 if len(clips_on_track) > 1 else None
                        ))

        plates.append(plate)

    plates.sort(key=lambda p: p.start_frame)
    return plates

def assign_episodes(plates, episode_markers):
    if not episode_markers:
        for plate in plates:
            plate.episode_code = "EP00"
        return

    # Sort markers by frame number to ensure correct order
    sorted_markers = sorted(episode_markers, key=lambda m: m.frame_number)

    for plate in plates:
        assigned = sorted_markers[0].episode_code
        for marker in sorted_markers:
            if plate.start_frame >= marker.frame_number:
                assigned = marker.episode_code
            # Don't break early - check all markers to find the last one that applies
        plate.episode_code = assigned

    counters = {}
    for plate in sorted(plates, key=lambda p: p.start_frame):
        ep = plate.episode_code
        counters[ep] = counters.get(ep, 0) + 1
        plate.plate_number = counters[ep]

def build_filename(show_code, show_acronym, date, episode_code, plate_number, layer=None):
    base = f"{show_code}_{show_acronym}_VFX_{episode_code}_{date}_{plate_number:02d}"
    if layer is None:
        return base
    if layer.layer_number == 0:
        return base + "_CLEAN"
    if layer.layer_number == -1:
        return base  # single orange with only CLEAN above — no V# suffix
    suffix = f"_V{layer.layer_number}"
    if layer.sub_index is not None:
        suffix += f"_{layer.sub_index:02d}"
    return base + suffix

def assign_filenames(plates, show_code, show_acronym, date):
    for plate in plates:
        if not plate.layers:
            clean = is_clip_apricot(plate.anchor_clip)
            plate.filename = build_filename(
                show_code, show_acronym, date, plate.episode_code, plate.plate_number,
                layer=PlateLayer(clip=None, track_index=2, start_frame=0, end_frame=0,
                                 layer_number=0) if clean else None)
        else:
            for layer in plate.layers:
                layer.filename = build_filename(
                    show_code, show_acronym, date, plate.episode_code, plate.plate_number, layer)

def get_export_list(plates):
    exports = []
    for plate in plates:
        if not plate.layers:
            exports.append({
                "clip": plate.anchor_clip,
                "track_index": 2,
                "filename": plate.filename,
                "episode_code": plate.episode_code,
                "plate_number": plate.plate_number,
                "start_frame": plate.start_frame,
                "end_frame": plate.end_frame,
            })
        else:
            for layer in plate.layers:
                exports.append({
                    "clip": layer.clip,
                    "track_index": layer.track_index,
                    "filename": layer.filename,
                    "episode_code": plate.episode_code,
                    "plate_number": plate.plate_number,
                    "start_frame": layer.start_frame,
                    "end_frame": layer.end_frame,
                })
    return exports

# ═══════════════════════════════════════════════════════════════════════════════
# RENDER
# ═══════════════════════════════════════════════════════════════════════════════

RENDER_PRESET = "02_COLORED VFX 4444 XQ"

def render_clip_now(project, timeline, clip, track_index, render_preset,
                    output_path, filename, mark_in, mark_out, log_callback=None):
    track_count = timeline.GetTrackCount("video")

    # Disable all tracks except the target track
    # First enable the target track in case it was disabled
    timeline.SetTrackEnable("video", track_index, True)
    
    disabled_tracks = []
    for t in range(1, track_count + 1):
        if t != track_index:
            timeline.SetTrackEnable("video", t, False)
            disabled_tracks.append(t)

    # Wait for DaVinci to process track changes
    time.sleep(0.5)
    
    # Verify
    active = [t for t in range(1, track_count + 1) if timeline.GetIsTrackEnabled("video", t)]
    if log_callback:
        log_callback(f"    Track state set: only track {track_index} should be active, got {active}")

    project.LoadRenderPreset(render_preset)
    project.SetRenderSettings({
        "SelectAllFrames": False,
        "MarkIn": mark_in,
        "MarkOut": mark_out,
        "TargetDir": output_path,
        "CustomName": filename,
        "UniqueFilenameStyle": 0,
        "ExportVideo": True,
        "ExportAudio": False,
    })

    clear_render_queue(project)
    job_id = project.AddRenderJob()

    if not job_id:
        if log_callback:
            log_callback(f"    Failed to add render job for {filename}")
        for t in disabled_tracks:
            timeline.SetTrackEnable("video", t, True)
        return False

    if log_callback:
        jobs = project.GetRenderJobList()
        for job in jobs:
            if job.get("JobId") == job_id:
                log_callback(f"    MarkIn={job.get('MarkIn')}, MarkOut={job.get('MarkOut')}, Track={track_index}")

    # Verify track state before rendering
    if log_callback:
        active = [t for t in range(1, track_count + 1) if timeline.GetIsTrackEnabled("video", t)]
        log_callback(f"    Active tracks before render: {active}")

    project.StartRendering(job_id)

    timeout = 600
    elapsed = 0
    while elapsed < timeout:
        time.sleep(1)
        elapsed += 1
        status = project.GetRenderJobStatus(job_id)
        job_status = status.get("JobStatus", "")
        if job_status == "Complete":
            break
        elif job_status == "Failed":
            if log_callback:
                log_callback(f"    Render failed: {status}")
            for t in disabled_tracks:
                timeline.SetTrackEnable("video", t, True)
            clear_render_queue(project)
            return False

    for t in disabled_tracks:
        timeline.SetTrackEnable("video", t, True)

    time.sleep(0.2)
    clear_render_queue(project)
    return True

def find_output_folder(show_code, show_acronym):
    """
    Auto-detect output folder by searching mounted volumes for SHOWCODE.
    Structure: /Volumes/SHOWCODE_*/SHOWCODE_*_EDIT/<folder with TURNOVER>/<folder with TO VFX>
    """
    import glob

    if not show_code:
        return None

    # Search /Volumes for drives starting with show_code
    volumes = glob.glob(f"/Volumes/{show_code}_*")
    if not volumes:
        return None

    for volume in sorted(volumes):
        # Look for SHOWCODE_*_EDIT folder inside
        edit_folders = glob.glob(f"{volume}/{show_code}_*_EDIT")
        if not edit_folders:
            edit_folders = glob.glob(f"{volume}/{show_code}_{show_acronym}_EDIT")
        if not edit_folders:
            continue
        for edit_folder in sorted(edit_folders):
            # Find folder with TURNOVER in name
            try:
                turnover_folders = [
                    os.path.join(edit_folder, d)
                    for d in os.listdir(edit_folder)
                    if "TURNOVER" in d.upper() and
                    os.path.isdir(os.path.join(edit_folder, d))
                ]
            except Exception:
                continue
            for turnover_folder in sorted(turnover_folders):
                # Find folder with TO VFX in name inside turnover folder
                try:
                    to_vfx_folders = [
                        os.path.join(turnover_folder, d)
                        for d in os.listdir(turnover_folder)
                        if "TO VFX" in d.upper() and
                        os.path.isdir(os.path.join(turnover_folder, d))
                    ]
                except Exception:
                    continue
                for to_vfx in sorted(to_vfx_folders):
                    return to_vfx

    return None


def episode_sort_key(ep_code):
    """Sort EP01 < EP01A < EP01B < EP02 < EP45A etc."""
    import re
    m = re.match('EP([0-9]+)([A-Z]?)', ep_code.upper())
    if m:
        return (int(m.group(1)), m.group(2))
    return (9999, "")


def generate_plate_list_xlsx(export_list, output_dir, show_code, acronym, date_str,
                              shot_map=None, list_folder=None, log_callback=None,
                              timeline=None, **kwargs):
    """Generate plate list xlsx using pre-captured screenshots from shot_map."""
    import openpyxl, os
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    if shot_map is None:
        shot_map = {}
    if list_folder is None:
        list_folder = os.path.join(output_dir, "LIST TO POST")
    os.makedirs(list_folder, exist_ok=True)

    filename = f"{show_code}_{acronym}_VFX_LIST TO POST_{date_str}.xlsx"
    out_path = os.path.join(list_folder, filename)

    # FPS and start TC for timecode display
    fps = 24.0
    start_tc_offset = "00:00:00:00"
    if timeline:
        try:
            fps = float(timeline.GetSetting("timelineFrameRate") or 24)
            start_tc_offset = timeline.GetStartTimecode() or "00:00:00:00"
        except:
            pass

    def frames_to_tc(frame_num):
        try:
            parts = start_tc_offset.split(":")
            offset = (int(parts[0])*3600 + int(parts[1])*60 + int(parts[2])) * int(fps) + int(parts[3])
            total = frame_num
        except:
            total = frame_num
        fi = int(fps)
        ff = total % fi; total //= fi
        ss = total % 60; total //= 60
        mm = total % 60; hh = total // 60
        return f"{hh:02d}:{mm:02d}:{ss:02d}:{ff:02d}"

    # Group export_list: one row per unique plate (V# layers share same start/end)
    seen = {}
    for item in export_list:
        key = (item["episode_code"], item["start_frame"], item["end_frame"])
        if key not in seen:
            seen[key] = {"item": item, "layers": 1}
        else:
            seen[key]["layers"] += 1
    unique_plates = sorted(seen.values(), key=lambda x: x["item"]["start_frame"])

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plate List"

    # Header style
    header_fill = PatternFill("solid", start_color="1E1E1E")
    header_font = Font(name="Arial", bold=True, color="E8A838", size=10)
    cell_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333")
    )

    columns = ["FILE NAME", "EPISODE", "PLATES TO TURNOVER",
               "VFX REF TIMECODE_IN", "VFX REF TIMECODE_OUT",
               "TURNOVER STATUS", "TURNOVER NOTE", "VFX", "VFX NOTE", "SCREENSHOT"]
    col_widths = [45, 12, 20, 22, 22, 18, 20, 12, 20, 25]

    for col_idx, (col_name, col_w) in enumerate(zip(columns, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_w

    ws.row_dimensions[1].height = 24

    # Data rows

    for row_idx, plate_data in enumerate(unique_plates, 2):
        item = plate_data["item"]
        layers = plate_data["layers"]

        tc_in = frames_to_tc(item["start_frame"])
        tc_out = frames_to_tc(item["end_frame"])

        values = [
            item["filename"] + ".mov",
            item["episode_code"],
            layers,
            tc_in,
            tc_out,
            "", "", "", "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = center if col_idx > 1 else left
            cell.border = thin


        # Screenshot — use pre-captured shot from shot_map
        shot_data = shot_map.get(item["filename"])
        if shot_data:
            shot_path, img_w, img_h = shot_data if isinstance(shot_data, tuple) else (shot_data, 160, 90)
            if os.path.exists(shot_path):
                # Row = 190px tall, image fills it exactly
                CELL_H_PX = 190
                ratio = CELL_H_PX / img_h
                disp_w = int(img_w * ratio)
                disp_h = CELL_H_PX
                xl_img = XLImage(shot_path)
                xl_img.width = disp_w
                xl_img.height = disp_h
                col_letter = openpyxl.utils.get_column_letter(10)
                ws.add_image(xl_img, f"{col_letter}{row_idx}")
                ws.row_dimensions[row_idx].height = CELL_H_PX * 72 / 96  # px to pt
                ws.column_dimensions[col_letter].width = max(
                    ws.column_dimensions[col_letter].width or 0, disp_w / 7 + 2)

        if log_callback:
            log_callback(f"  [{row_idx-1}/{len(unique_plates)}] {item['filename']}")

    wb.save(out_path)

    if log_callback:
        log_callback(f"✓ Plate list saved: {filename}")

    return out_path


def find_last_exported(output_dir, export_list):
    """
    Scan output folder to find which clips are already exported.
    Returns the index in export_list to resume from.
    """
    import glob
    exported = set()
    for mov in glob.glob(os.path.join(output_dir, "EP*", "*.mov")):
        exported.add(os.path.basename(mov).replace(".mov", ""))

    if not exported:
        return 0

    # Find the last consecutive index that was exported
    # (not just the last one found, but the last one in sequence)
    last_done = -1
    for i, item in enumerate(export_list):
        if item["filename"] in exported:
            last_done = i
        else:
            # First gap found - resume from here
            if last_done >= 0:
                return last_done + 1

    # All found or none found

    return last_done + 1  # resume from next clip


def render_single_clip(project, timeline, render_preset,
                       output_path, filename, mark_in, mark_out, log_callback=None):
    """Render one clip — tracks must already be set correctly before calling."""
    project.LoadRenderPreset(render_preset)
    project.SetRenderSettings({
        "SelectAllFrames": False,
        "MarkIn": mark_in,
        "MarkOut": mark_out,
        "TargetDir": output_path,
        "CustomName": filename,
        "UniqueFilenameStyle": 0,
        "ExportVideo": True,
        "ExportAudio": False,
    })

    clear_render_queue(project)
    job_id = project.AddRenderJob()

    if not job_id:
        if log_callback:
            log_callback(f"    Failed to add render job for {filename}")
        return False

    if log_callback:
        jobs = project.GetRenderJobList()
        for job in jobs:
            if job.get("JobId") == job_id:
                log_callback(f"    MarkIn={job.get('MarkIn')}, MarkOut={job.get('MarkOut')}")

    project.StartRendering(job_id)

    timeout = 600
    elapsed = 0
    while elapsed < timeout:
        time.sleep(1)
        elapsed += 1
        status = project.GetRenderJobStatus(job_id)
        job_status = status.get("JobStatus", "")
        if job_status == "Complete":
            break
        elif job_status == "Failed":
            if log_callback:
                log_callback(f"    Render failed: {status}")
            clear_render_queue(project)
            return False

    clear_render_queue(project)
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class ExportEngine:
    def __init__(self, log_callback=None, progress_callback=None):
        self.log = log_callback or print
        self.progress = progress_callback or (lambda pct, msg: None)
        self.resolve = None
        self.project = None
        self.timeline = None
        self.episode_markers = []
        self.export_list = []

    def connect(self):
        self.log("Connecting to DaVinci Resolve...")
        self.resolve = get_resolve()
        self.project = get_current_project(self.resolve)
        self.timeline = get_current_timeline(self.project)
        info = get_timeline_info(self.timeline)
        self.log(f"Connected. Timeline: {info['name']} @ {info['frame_rate']} fps")

        parts = info["name"].split("_")
        info["show_code"] = parts[0] if len(parts) > 0 else ""
        info["show_acronym"] = parts[1] if len(parts) > 1 else ""

        start_tc = self.timeline.GetStartTimecode()
        info["start_timecode"] = start_tc
        info["start_frame"] = self.timeline.GetStartFrame()
        info["start_hour"] = int(start_tc.split(":")[0]) if start_tc else 3
        self.log(f"Timeline start timecode: {start_tc}")

        return info

    def scan_episodes(self, reference_video_path=None, start_timecode="03:59:50:00", stop_flag=None):
        import cv2, time

        if not reference_video_path:
            self.log("Auto-detecting reference video from project...")
            reference_video_path = get_offline_reference_path(self.project)
        if not reference_video_path or not os.path.exists(reference_video_path):
            raise FileNotFoundError("Could not find reference video. Please select it manually.")

        self.log(f"Reference video: {os.path.basename(reference_video_path)}")

        info = get_timeline_info(self.timeline)
        frame_rate = float(info["frame_rate"])

        self.log("Finding gaps across all tracks...")
        track_count = get_track_count(self.timeline, "video")
        all_ranges = []
        for t in range(1, track_count + 1):
            clips = get_all_clips_on_track(self.timeline, "video", t)
            for c in clips:
                all_ranges.append((get_clip_start_frame(c), get_clip_end_frame(c)))

        if not all_ranges:
            self.log("No clips found in timeline.")
            self.episode_markers = []
            return self.episode_markers

        all_ranges.sort(key=lambda x: x[0])
        merged = [list(all_ranges[0])]
        for start, end in all_ranges[1:]:
            if start <= merged[-1][1]:
                merged[-1][1] = max(merged[-1][1], end)
            else:
                merged.append([start, end])

        gaps = []
        timeline_start = self.timeline.GetStartFrame()
        if merged[0][0] > timeline_start:
            gap_len = merged[0][0] - timeline_start
            if gap_len >= 20:
                gaps.append((timeline_start, merged[0][0]))
        for i in range(len(merged) - 1):
            gap_start = merged[i][1]
            gap_end = merged[i + 1][0]
            if gap_end - gap_start >= 20:
                gaps.append((gap_start, gap_end))

        self.log(f"Found {len(gaps)} gaps ≥ 20 frames in timeline")

        if not gaps:
            self.log("No gaps found.")
            self.episode_markers = []
            return self.episode_markers

        # Step 1: Get timecodes from DaVinci for each gap
        self.log("Getting timecodes from DaVinci...")
        gap_timecodes = []
        for i, (gap_start, gap_end) in enumerate(gaps):
            if stop_flag and stop_flag():
                self.log("Scan stopped by user.")
                cap_ref = None
                break
            pct = (i / len(gaps)) * 40
            self.progress(pct, f"Getting timecode {i+1}/{len(gaps)}...")

            if i == 0:
                mid_frame = max(gap_end - 24, gap_start + 1)
            else:
                mid_frame = min(((gap_start + gap_end) // 2) + 10, gap_end - 1)

            target_tc = frames_to_davinci_tc(
                mid_frame, frame_rate,
                self.timeline.GetStartFrame(),
                self.timeline.GetStartTimecode()
            )
            self.timeline.SetCurrentTimecode(target_tc)
            confirmed_tc = self.timeline.GetCurrentTimecode()
            gap_timecodes.append((confirmed_tc, gaps[i][0]))
            self.log(f"  Gap {i+1}: {confirmed_tc}")

        self.resolve.OpenPage("edit")
        self.log("Returned to Edit page. Now scanning reference video...")

        # Step 2: Seek into reference video and OCR
        video_start_frames = timecode_str_to_frames(start_timecode, frame_rate)

        cap = cv2.VideoCapture(reference_video_path)
        if not cap.isOpened():
            raise RuntimeError(f"Could not open reference video: {reference_video_path}")

        markers = []
        last_episode = None

        for i, (tc_str, gap_start_frame) in enumerate(gap_timecodes):
            if stop_flag and stop_flag():
                self.log("Scan stopped by user.")
                break
            pct = 40 + (i / len(gap_timecodes)) * 50
            self.progress(pct, f"OCR on gap {i+1}/{len(gap_timecodes)}...")

            parts = tc_str.split(":")
            hh, mm, ss, ff = int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3])
            fps = round(float(frame_rate))
            tc_total_frames = ((hh * 3600 + mm * 60 + ss) * fps) + ff
            file_frame = tc_total_frames - video_start_frames

            if file_frame < 0:
                continue

            cap.set(cv2.CAP_PROP_POS_FRAMES, file_frame)
            ret, frame = cap.read()
            if not ret or frame is None:
                self.log(f"  Gap {i+1}: could not read frame")
                continue

            code = extract_episode_code(frame)
            self.log(f"  Gap {i+1} at {tc_str}: OCR={repr(code)}")

            if code and code != last_episode:
                markers.append(EpisodeMarker(
                    episode_code=code,
                    frame_number=gap_start_frame,
                    timecode=tc_str
                ))
                last_episode = code
                self.log(f"  ✓ Found {code} at {tc_str}")

        cap.release()

        self.episode_markers = markers
        self.progress(95, "Scan complete.")
        self.log(f"Found {len(self.episode_markers)} episodes:")
        for m in self.episode_markers:
            self.log(f"  {m.episode_code} → {m.timecode}")
        return self.episode_markers

    def prepare_export(self, show_code, show_acronym, date):
        self.log("Scanning for orange VFX plates...")
        self.progress(60, "Scanning for VFX plates...")
        plates = get_vfx_plates(self.timeline)
        self.log(f"Found {len(plates)} VFX plates")
        for p in plates:
            if p.layers:
                self.log(f"  Plate at {p.start_frame}-{p.end_frame}: {len(p.layers)} layers: {[l.layer_number for l in p.layers]}")
            else:
                self.log(f"  Plate at {p.start_frame}-{p.end_frame}: no layers")
        self.progress(70, "Assigning episode codes...")
        assign_episodes(plates, self.episode_markers)
        self.progress(80, "Building filenames...")
        assign_filenames(plates, show_code, show_acronym, date)
        self.export_list = get_export_list(plates)
        self.log(f"\nReady to export {len(self.export_list)} clips:")
        for item in self.export_list:
            self.log(f"  {item['filename']}")
        return self.export_list

    def run_export(self, output_dir, stop_flag=None, start_index=0, screenshot_callback=None):
        if not self.export_list:
            raise RuntimeError("No clips to export. Run prepare_export first.")
        os.makedirs(output_dir, exist_ok=True)

        # Create episode folders and 01_ARCHIVE subfolders
        episodes = set(item["episode_code"] for item in self.export_list)
        for ep in sorted(episodes):
            ep_folder = os.path.join(output_dir, ep)
            os.makedirs(ep_folder, exist_ok=True)
            os.makedirs(os.path.join(ep_folder, "01_ARCHIVE"), exist_ok=True)
            self.log(f"  Created folder: {ep}/")

        total = len(self.export_list)
        track_count = self.timeline.GetTrackCount("video")
        self.log(f"\nExporting {total} clips track by track...")

        # Group clips by track, preserving order
        from collections import defaultdict
        clips_by_track = defaultdict(list)
        for item in self.export_list:
            clips_by_track[item["track_index"]].append(item)

        done = start_index  # count already-done clips in progress

        # Filter export list to only clips from start_index onward
        remaining = self.export_list[start_index:]
        clips_by_track = defaultdict(list)
        for item in remaining:
            clips_by_track[item["track_index"]].append(item)

        for track_index in sorted(clips_by_track.keys()):
            track_clips = clips_by_track[track_index]
            self.log(f"\n── Track {track_index}: {len(track_clips)} clips ──")

            # ── Edit page: mute all tracks except this one ─────────────────
            self.resolve.OpenPage("edit")
            time.sleep(1.5)

            for t in range(1, track_count + 1):
                self.timeline.SetTrackEnable("video", t, t == track_index)
            time.sleep(1)

            active = [t for t in range(1, track_count + 1)
                     if self.timeline.GetIsTrackEnabled("video", t)]
            self.log(f"  Active tracks: {active}")

            # ── Deliver page: render all clips on this track ───────────────
            self.resolve.OpenPage("deliver")
            time.sleep(1.5)

            for item in track_clips:
                if stop_flag and stop_flag():
                    self.log("Export stopped by user.")
                    # Clean up DaVinci state so it's ready for continue
                    clear_render_queue(self.project)
                    self.resolve.OpenPage("edit")
                    time.sleep(0.5)
                    for t in range(1, track_count + 1):
                        self.timeline.SetTrackEnable("video", t, True)
                    return False

                done += 1
                pct = (done / total) * 100
                track_num = item.get('track_index', '?')
                self.progress(pct, f"Exporting Plate: {done}/{total} · {item['episode_code']} · Track {track_num}")
                self.log(f"  [{done}/{total}] {item['filename']} · Track {track_num}")

                # Grab screenshot before rendering
                if screenshot_callback:
                    screenshot_callback(item)

                ep_folder = os.path.join(output_dir, item["episode_code"])
                success = render_single_clip(
                    project=self.project,
                    timeline=self.timeline,
                    render_preset=RENDER_PRESET,
                    output_path=ep_folder,
                    filename=item["filename"],
                    mark_in=item["start_frame"],
                    mark_out=item["end_frame"] - 1,
                    log_callback=self.log
                )

                if not success:
                    self.log(f"  ✗ Failed: {item['filename']}")
                else:
                    self.log(f"  ✓ Done: {item['filename']}")
                    # Only update _last_done after confirmed success
                    self._last_done = done

        # ── Restore all tracks ─────────────────────────────────────────────
        self.resolve.OpenPage("edit")
        time.sleep(1)
        self.log("\nRestoring all tracks...")
        for t in range(1, track_count + 1):
            self.timeline.SetTrackEnable("video", t, True)

        self.progress(100, "Export complete!")
        self.log(f"\nAll {total} clips exported to: {output_dir}")
        return True

# ═══════════════════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════════════════

BG_DARK      = "#1E1E1E"
BG_OUTER     = "#161616"  # slightly darker for outer app background
BG_PANEL     = "#2A2A2A"
BG_INPUT     = "#333333"
ACCENT       = "#E8A838"
ACCENT_HOVER = "#F0BC5A"
TEXT_PRIMARY = "#F0F0F0"
TEXT_MUTED   = "#888888"
TEXT_SUCCESS = "#5DBD74"
TEXT_ERROR   = "#E05555"
TEXT_WARN    = "#E8A838"
BORDER       = "#3A3A3A"
FONT_MAIN    = ("SF Pro Display", 13)
FONT_SMALL   = ("SF Pro Display", 11)
FONT_MONO    = ("SF Mono", 11)
FONT_TITLE   = ("SF Pro Display", 22, "bold")
FONT_LABEL   = ("SF Pro Display", 12)


class VFXExporterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Finishing Tool")
        self.configure(bg=BG_OUTER)
        try:
            from ctypes import windll
        except ImportError:
            pass
        # Fix Retina/HiDPI scaling when launched as .app bundle
        try:
            self.tk.call("tk", "scaling", 1.0)
        except Exception:
            pass
        self.resizable(True, True)
        self.minsize(860, 750)

        self.engine = None
        self.episode_markers = []
        self.export_list = []
        self.reference_video_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.expanduser("~/Desktop/VFX_Export"))
        self.show_code = tk.StringVar()
        self.show_acronym = tk.StringVar()
        self.export_date = tk.StringVar(value=datetime.now().strftime("%y%m%d"))
        self.start_timecode = tk.StringVar(value="")
        self.start_timecode_display = tk.StringVar(value="Auto-detected on connect")
        self._stop_export = False
        self._disabled_episodes = set()
        self._export_started = False  # episodes toggled off by user
        self._stop_scan = False
        self._scanning = False
        self._suppress_progress = False
        self.filename_preview = tk.StringVar(value="Connect to DaVinci to auto-fill...")

        # Independent mode vars for each section
        self.show_mode = tk.StringVar(value="auto")
        self.ref_mode = tk.StringVar(value="auto")
        self.out_mode = tk.StringVar(value="auto")

        for var in [self.show_code, self.show_acronym, self.export_date]:
            var.trace_add("write", lambda *args: self.after(0, self._update_filename_preview))

        self._build_ui()
        self.after(200, self._disable_reset_btn)
        self._check_deps_on_start()
        self.after(2000, self._check_for_updates)

    def _build_ui(self):
        main = tk.Frame(self, bg=BG_OUTER, padx=28, pady=24)
        main.pack(fill="both", expand=True)
        self._main_frame = main

        # ── App title + Reset button ───────────────────────────────────────
        title_row = tk.Frame(main, bg=BG_OUTER)
        title_row.pack(fill="x", pady=(0, 0))
        title_lbl_frame = tk.Frame(title_row, bg=BG_OUTER)
        title_lbl_frame.pack(side="left")
        tk.Label(title_lbl_frame, text="Finishing Tool", font=FONT_TITLE,
                 bg=BG_OUTER, fg=ACCENT).pack(side="left")
        self._version_label = tk.Label(title_lbl_frame, text=f"v{APP_VERSION}", font=("SF Pro Display", 13),
                 bg=BG_OUTER, fg=TEXT_MUTED)
        self._version_label.pack(side="left", padx=(8, 0), pady=(6, 0))
        # Reset button as rounded canvas
        # Reset button as canvas for color control on macOS
        def _make_reset_canvas(text, bg, fg):
            font = ("SF Pro Display", 12, "bold")
            import tkinter.font as tkfont
            f = tkfont.Font(family="SF Pro Display", size=12, weight="bold")
            tw, th = f.measure(text), f.metrics("linespace")
            w, h, r = tw + 32, th + 14, 7
            c = tk.Canvas(title_row, width=w, height=h, bg=BG_OUTER,
                          highlightthickness=0, cursor="")
            x1,y1,x2,y2 = 0,0,w,h
            pts = [x1+r,y1,x2-r,y1,x2,y1,x2,y1+r,x2,y2-r,x2,y2,
                   x2-r,y2,x1+r,y2,x1,y2,x1,y2-r,x1,y1+r,x1,y1]
            poly_id = c.create_polygon(pts, fill=bg, outline="", smooth=True)
            text_id = c.create_text(w//2, h//2, font=font, fill=fg)
            c._text = text
            c._bg = bg
            c.itemconfig(text_id, text=text)

            def draw(fill):
                c.itemconfig(poly_id, fill=fill)
                c.itemconfig(text_id, text=c._text)

            c._draw = draw
            c.bind("<Enter>", lambda e: draw("#3a8e3a" if c._bg=="#2a6e2a" else "#ae3a3a"))
            c.bind("<Leave>", lambda e: draw(c._bg))
            c.bind("<ButtonPress-1>", lambda e: draw("#1a5e1a" if c._bg=="#2a6e2a" else "#6e1a1a"))
            # ButtonRelease handled by _enable_reset_btn bindings
            pass
            c._action = lambda: None
            return c

        self.btn_reset = _make_reset_canvas("RESET ALL", "#2a6e2a", "#FFFFFF")
        self.btn_reset.pack(side="right")
        self.btn_reset._action = self._do_reset
        self.btn_reset._enabled = False

        tk.Label(main, text="DaVinci Resolve API",
                 font=FONT_SMALL, bg=BG_OUTER, fg=TEXT_MUTED).pack(anchor="w", pady=(0, 36))
        self._thinking_active = False
        self._thinking_frames = []
        self._thinking_frame_idx = 0

        # Load GIF frames
        import os as _os
        # Find GIF next to the script, with fallback paths
        _script_dir = _os.path.dirname(_os.path.abspath(__file__)) if "__file__" in dir() else _os.path.expanduser("~/Downloads/vfx_exporter")
        gif_path = _os.path.join(_script_dir, "thinking.gif")
        try:
            from PIL import Image, ImageTk
            if not _os.path.exists(gif_path):
                raise FileNotFoundError(f"GIF not found: {gif_path}")
            gif = Image.open(gif_path)
            while True:
                frame = gif.copy().convert("RGB")
                self._thinking_frames.append(ImageTk.PhotoImage(frame))
                try:
                    gif.seek(gif.tell() + 1)
                except EOFError:
                    break
        except Exception:
            self._thinking_frames = []

        # ── Tab bar ────────────────────────────────────────────────────────
        self._active_tab = "vfx"
        self._tab_canvases = {}
        tab_bar = tk.Frame(main, bg=BG_OUTER)
        tab_bar.pack(fill="x", pady=(0, 0))
        # Thinking indicator on right side of tab bar
        self._thinking_label = tk.Label(tab_bar, bg=BG_OUTER)
        self._thinking_label.pack(side="right", padx=(0, 4))
        # Status text for connect (shown instead of GIF during connect)
        self._connect_status = tk.Label(tab_bar, text="", font=FONT_SMALL,
                                         bg=BG_OUTER, fg=TEXT_MUTED)
        self._connect_status.pack(side="right", padx=(0, 8))

        def _draw_tab(canvas, label, active):
            tw, th, r = 130, 36, 10
            col = ACCENT if active else "#888888"
            bg = BG_DARK if active else "#252525"

            lw = 1.5 if active else 1
            canvas.delete("all")
            canvas.create_rectangle(0, r, tw, th+1, fill=bg, outline="")
            canvas.create_rectangle(r, 0, tw-r, r, fill=bg, outline="")
            canvas.create_arc(0, 0, r*2, r*2, start=90, extent=90, fill=bg, outline="")
            canvas.create_arc(tw-r*2, 0, tw, r*2, start=0, extent=90, fill=bg, outline="")
            if active:
                canvas.create_line(r, 1, tw-r, 1, fill=col, width=lw)
                canvas.create_arc(1, 1, r*2+1, r*2+1, start=90, extent=90,
                                  outline=col, style="arc", width=lw)
                canvas.create_arc(tw-r*2-1, 1, tw-1, r*2+1, start=0, extent=90,
                                  outline=col, style="arc", width=lw)
                canvas.create_line(1, r, 1, th+1, fill=col, width=lw)
                canvas.create_line(tw-1, r, tw-1, th+1, fill=col, width=lw)
            fw = "bold" if active else "normal"
            canvas.create_text(tw//2, th//2, text=label,
                              font=("SF Pro Display", 12, fw), fill=col)

        def _make_tab(label, tab_id):
            tw, th = 130, 36
            c = tk.Canvas(tab_bar, width=tw, height=th+1, bg=BG_OUTER, highlightthickness=0)
            c.pack(side="left", padx=(0, 2))
            self._tab_canvases[tab_id] = (c, label)
            _draw_tab(c, label, tab_id == self._active_tab)
            c.bind("<Button-1>", lambda e, tid=tab_id: self._switch_tab(tid))

        _make_tab("VFX Export", "vfx")
        _make_tab("TEST", "test")
        self._draw_tab_fn = lambda: [_draw_tab(c, l, tid == self._active_tab)
                                      for tid, (c, l) in self._tab_canvases.items()]

        # ── Tab content panel ──────────────────────────────────────────────
        self.tab_content_frame = tk.Frame(main, bg=BG_DARK,
                                           highlightthickness=1,
                                           highlightbackground="#5a5a5a",
                                           highlightcolor="#5a5a5a")
        self.tab_content_frame.pack(fill="both", expand=True, pady=(0, 8))

        self._vfx_content = tk.Frame(self.tab_content_frame, bg=BG_DARK, padx=16, pady=12)
        self._vfx_content.pack(fill="both", expand=True)
        self.tab_content = self._vfx_content
        self._build_vfx_tab(self._vfx_content)

        self._test_content = tk.Frame(self.tab_content_frame, bg=BG_DARK, padx=16, pady=12)
        self._section_label(self._test_content, "TEST")
        test_panel = self._panel(self._test_content)
        tk.Label(test_panel, text="Hello", font=FONT_MAIN, bg=BG_PANEL,
                 fg=TEXT_PRIMARY).pack(anchor="w")

    def _build_vfx_tab(self, main):
        # ── SHOW INFO ──────────────────────────────────────────────────────
        self._section_label(main, "SHOW INFO")
        show_frame = self._panel(main)

        self._mode_row(show_frame, self.show_mode, self._on_show_mode_change)

        prev_row = tk.Frame(show_frame, bg=BG_PANEL)
        prev_row.pack(fill="x", pady=4)
        tk.Label(prev_row, text="Filename Preview", font=FONT_LABEL, bg=BG_PANEL,
                 fg=TEXT_PRIMARY, width=22, anchor="w").pack(side="left")
        self.preview_label = tk.Label(prev_row, textvariable=self.filename_preview,
                                       font=FONT_SMALL, bg=BG_PANEL, fg=ACCENT, anchor="w")
        self.preview_label.pack(side="left", fill="x", expand=True)

        self.show_manual_frame = tk.Frame(show_frame, bg=BG_PANEL)
        self._field_row(self.show_manual_frame, "Show Code", self.show_code, "e.g. V-LA35")
        self._field_row(self.show_manual_frame, "Show Acronym", self.show_acronym, "e.g. HBFSHR")
        self._field_row(self.show_manual_frame, "Export Date (YYMMDD)", self.export_date, "e.g. 260617")

        tc_row = tk.Frame(show_frame, bg=BG_PANEL)
        tc_row.pack(fill="x", pady=4)
        tk.Label(tc_row, text="Ref Video Start TC", font=FONT_LABEL, bg=BG_PANEL,
                 fg=TEXT_PRIMARY, width=22, anchor="w").pack(side="left")
        self.hour_display = tk.Label(tc_row, textvariable=self.start_timecode_display,
                                      font=FONT_SMALL, bg=BG_PANEL, fg=TEXT_MUTED)
        self.hour_display.pack(side="left")
        self.hour_wrap = tk.Frame(tc_row, bg=BG_INPUT,
                             highlightthickness=1,
                             highlightbackground=BORDER, highlightcolor=ACCENT)
        tk.Frame(self.hour_wrap, bg=BG_INPUT, width=2).pack(side="left")
        self.hour_entry = tk.Entry(self.hour_wrap, textvariable=self.start_timecode,
                                    font=FONT_SMALL, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                    insertbackground=TEXT_PRIMARY, relief="flat", bd=0,
                                    highlightthickness=0, width=10)
        self.hour_entry.pack(side="left", ipady=7)
        # hour_wrap shown in manual mode only; hour_display shown in auto mode (default)
        self.hour_hint = tk.Label(tc_row, text="hh:mm:ss:ff",
                                   font=FONT_SMALL, bg=BG_PANEL, fg="#666666")

        # ── REFERENCE VIDEO ────────────────────────────────────────────────
        self._section_label(main, "REFERENCE VIDEO")
        ref_frame = self._panel(main)
        self._mode_row(ref_frame, self.ref_mode, self._on_ref_mode_change)
        ref_row = tk.Frame(ref_frame, bg=BG_PANEL)
        ref_row.pack(fill="x", pady=4)
        ref_wrap = tk.Frame(ref_row, bg=BG_INPUT,
                            highlightthickness=1,
                            highlightbackground=BORDER, highlightcolor=ACCENT)
        ref_wrap.pack(side="left", fill="x", expand=True)
        tk.Frame(ref_wrap, bg=BG_INPUT, width=2).pack(side="left")
        self.ref_entry = tk.Entry(ref_wrap, textvariable=self.reference_video_path,
                 font=FONT_SMALL, bg=BG_INPUT, fg=TEXT_PRIMARY,
                 insertbackground=TEXT_PRIMARY, relief="flat", bd=0, highlightthickness=0)
        self.ref_entry.pack(side="left", fill="x", expand=True, ipady=7)
        self.ref_browse_btn = self._rounded_btn(ref_row, "Browse...", self._browse_reference, small=True, match_height=True)
        self.ref_browse_btn.pack(side="left", padx=(8, 0), fill="y")
        self.ref_hint = tk.Label(ref_frame, text="Auto-detected on connect",
                 font=FONT_SMALL, bg=BG_PANEL, fg=TEXT_MUTED)
        self.ref_hint.pack(anchor="w", pady=(4, 0))

        # ── OUTPUT FOLDER ──────────────────────────────────────────────────
        self._section_label(main, "OUTPUT FOLDER")
        out_frame = self._panel(main)
        self._mode_row(out_frame, self.out_mode, self._on_out_mode_change)
        out_row = tk.Frame(out_frame, bg=BG_PANEL)
        out_row.pack(fill="x", pady=4)
        out_wrap = tk.Frame(out_row, bg=BG_INPUT,
                            highlightthickness=1,
                            highlightbackground=BORDER, highlightcolor=ACCENT)
        out_wrap.pack(side="left", fill="x", expand=True)
        tk.Frame(out_wrap, bg=BG_INPUT, width=2).pack(side="left")
        self.out_entry = tk.Entry(out_wrap, textvariable=self.output_dir,
                 font=FONT_SMALL, bg=BG_INPUT, fg=TEXT_PRIMARY,
                 insertbackground=TEXT_PRIMARY, relief="flat", bd=0, highlightthickness=0)
        self.out_entry.pack(side="left", fill="x", expand=True, ipady=7)
        self.out_browse_btn = self._rounded_btn(out_row, "Browse...", self._browse_output, small=True, match_height=True)
        self.out_browse_btn.pack(side="left", padx=(8, 0), fill="y")
        self.out_hint = tk.Label(out_frame, text="Auto-detected on connect",
                 font=FONT_SMALL, bg=BG_PANEL, fg=TEXT_MUTED)
        self.out_hint.pack(anchor="w", pady=(4, 0))

        # Initialize widget states based on current mode (both default to auto = disabled)
        # These are called at end of _build_vfx_tab after all widgets exist

        # ── EXPORT BUTTONS (3 steps only) ──────────────────────────────────
        self._section_label(main, "EXPORT")
        btn_frame = tk.Frame(main, bg=BG_DARK)
        btn_frame.pack(fill="x", pady=(0, 12), padx=(4, 0))
        self.step_canvases = []

        def _num(parent, n):
            size = 24
            c = tk.Canvas(parent, width=size, height=size, bg=BG_DARK, highlightthickness=0)
            c.pack(side="left", padx=(0, 8))
            c.create_oval(1, 1, size-1, size-1, outline=ACCENT, width=2, fill="")
            c.create_text(size//2, size//2, text=str(n),
                         font=("SF Pro Display", 12, "bold"), fill=ACCENT)
            self.step_canvases.append(c)

        _num(btn_frame, "1")
        self.btn_connect = self._rounded_btn(btn_frame, "Connect to Resolve", self._do_connect)
        self.btn_connect.pack(side="left", padx=(0, 14))
        _num(btn_frame, "2")
        self.btn_scan = self._rounded_btn(btn_frame, "Scan Episodes", self._do_scan, enabled=False)
        self.btn_scan.pack(side="left", padx=(0, 14))
        _num(btn_frame, "3")
        self.btn_export = self._rounded_btn(btn_frame, "Export", self._do_export,
                                              enabled=False, accent=True,
                                              reserve_text="Continue")
        self.btn_export.pack(side="left", padx=(0, 8))
        # Reset Export sits right next to Export button, always visible but starts disabled
        self.btn_reset_export = self._rounded_btn(btn_frame, "Reset Export",
                                                   self._do_reset_export, enabled=False)
        self.btn_reset_export.pack(side="left", padx=(8, 0))
        # Starts disabled — enabled via _set_btn_state when export stops or completes

        # Create CSV checkbox — right of Reset Export, disabled until scan done
        self._create_xlsx = tk.BooleanVar(value=True)
        self._xlsx_check = tk.Checkbutton(btn_frame, text="Create .xlsx",
                                           variable=self._create_xlsx,
                                           font=("SF Pro Display", 11),
                                           bg=BG_DARK, fg=TEXT_MUTED,
                                           selectcolor=BG_DARK,
                                           activebackground=BG_DARK,
                                           activeforeground=TEXT_PRIMARY,
                                           disabledforeground="#444444",
                                           state="normal", cursor="")
        self._xlsx_check.pack(side="left", padx=(16, 0))

        # ── PROGRESS ───────────────────────────────────────────────────────
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(main, variable=self.progress_var,
                                             maximum=100, mode="determinate")
        prog_style = ttk.Style()
        prog_style.theme_use("clam")
        prog_style.configure("TProgressbar", troughcolor=BG_INPUT,
                        background=ACCENT, borderwidth=0, thickness=10)
        self.progress_bar = ttk.Progressbar(main, variable=self.progress_var,
                                             maximum=100, mode="determinate")
        self.progress_bar.pack(fill="x", pady=(16, 4))
        self.progress_status = tk.Label(main, text="", font=FONT_SMALL,
                                         bg=BG_DARK, fg=TEXT_SUCCESS)
        self.progress_status.pack(anchor="w", pady=(0, 4))

        # (action buttons are in btn_frame, hidden until needed)

        # ── DETECTED EPISODES (tag style) ──────────────────────────────────
        ep_header = tk.Frame(main, bg=BG_DARK)
        ep_header.pack(fill="x", pady=(12, 0))
        tk.Label(ep_header, text="DETECTED EPISODES",
                 font=("SF Pro Display", 10, "bold"), bg=BG_DARK, fg=TEXT_MUTED).pack(side="left")
        self.btn_toggle_all = tk.Label(ep_header, text="DISABLE ALL",
                                        font=("SF Pro Display", 10, "bold"),
                                        bg=BG_INPUT, fg=TEXT_PRIMARY,
                                        padx=10, pady=3, cursor="")
        self.btn_toggle_all.pack(side="right")
        self.btn_toggle_all.bind("<Button-1>", lambda e: self._toggle_all_episodes())
        self._all_disabled = False
        ep_outer = tk.Frame(main, bg="#252525", padx=12, pady=12)
        ep_outer.pack(fill="x", pady=(0, 8))
        self.ep_tags_frame = tk.Frame(ep_outer, bg="#252525")
        self.ep_tags_frame.pack(fill="x")
        self._ep_tag_widgets = []

        # Initialize all mode states properly
        self._on_show_mode_change()
        self._on_ref_mode_change()
        self._on_out_mode_change()

        # ── NUMBER OF PLATES ───────────────────────────────────────────────
        self._section_label(main, "NUMBER OF PLATES DETECTED")
        plates_frame = tk.Frame(main, bg="#252525", padx=16, pady=12)
        plates_frame.pack(fill="x", pady=(0, 8))
        self.plates_count_label = tk.Label(plates_frame, text="—",
                 font=("SF Pro Display", 14, "bold"), bg="#252525", fg=ACCENT,
                 anchor="w")
        self.plates_count_label.pack(anchor="w")

        # ── LOG (hidden by default) ─────────────────────────────────────────
        log_header = tk.Frame(main, bg=BG_DARK)
        log_header.pack(fill="x", pady=(28, 0))
        self.btn_log_toggle = tk.Label(log_header, text="SHOW LOG",
                                        font=("SF Pro Display", 10, "bold"),
                                        bg=BG_INPUT, fg=TEXT_PRIMARY,
                                        padx=10, pady=3, cursor="")
        self.btn_log_toggle.pack(side="left")
        self.btn_log_toggle.bind("<Button-1>", lambda e: self._toggle_log())

        self.log_frame_outer = tk.Frame(main, bg=BG_DARK)
        log_p = self._panel(self.log_frame_outer)
        log_scroll = tk.Scrollbar(log_p, bg=BG_PANEL)
        log_scroll.pack(side="right", fill="y")
        self.log_box = tk.Text(log_p, font=FONT_MONO, bg=BG_INPUT, fg=TEXT_PRIMARY,
                                relief="flat", bd=0, height=16, width=1,
                                yscrollcommand=log_scroll.set, state="disabled", wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=8, pady=8)
        log_scroll.config(command=self.log_box.yview)
        self.log_box.tag_configure("success", foreground=TEXT_SUCCESS)
        self.log_box.tag_configure("error", foreground=TEXT_ERROR)
        self.log_box.tag_configure("warn", foreground=TEXT_WARN)
        self.log_box.tag_configure("muted", foreground=TEXT_MUTED)
        self._log_visible = False

    def _do_continue_export(self):
        output = self.output_dir.get().strip()
        if not output:
            return
        # Use _last_done from engine - exact index of last successfully exported clip
        last = getattr(self.engine, '_last_done', 0)
        start_index = last  # _last_done = number of clips successfully completed
        if start_index >= len(self.export_list):
            self._log("All clips already exported.", "success")
            return
        self._log(f"Resuming from clip {start_index + 1}/{len(self.export_list)}.", "muted")
        self.btn_export._text = "Export"
        self._set_btn_state(self.btn_export, False)
        self._set_stop_button()  # Ensure STOP shows immediately before thread starts
        self._run_export_task(output, start_index=start_index)

    def _enable_reset_export_btn(self):
        """Enable Reset Export button with green styling."""
        self.btn_reset_export._bg = "#2a6e2a"
        self.btn_reset_export._state["bg"] = "#2a6e2a"
        self.btn_reset_export._state["fg"] = "#FFFFFF"
        self.btn_reset_export._draw("#2a6e2a", "#FFFFFF")
        self.btn_reset_export.unbind("<Enter>")
        self.btn_reset_export.unbind("<Leave>")
        self.btn_reset_export.unbind("<ButtonPress-1>")
        self.btn_reset_export.unbind("<ButtonRelease-1>")
        self.btn_reset_export.bind("<Enter>", lambda e: self.btn_reset_export._draw("#3a8e3a", "#FFFFFF"))
        self.btn_reset_export.bind("<Leave>", lambda e: self.btn_reset_export._draw("#2a6e2a", "#FFFFFF"))
        self.btn_reset_export.bind("<ButtonPress-1>", lambda e: self.btn_reset_export._draw("#1a5e1a", "#FFFFFF"))
        self.btn_reset_export.bind("<ButtonRelease-1>", lambda e: (self._do_reset_export(), self.btn_reset_export._draw("#2a6e2a", "#FFFFFF")))

    def _do_reset_export(self):
        """Reset export progress but keep episodes. Restore full plate list."""
        self._set_btn_state(self.btn_reset_export, False)
        self._export_started = False
        self._all_disabled = False
        self.btn_toggle_all.config(text="DISABLE ALL", bg=BG_INPUT, fg=TEXT_PRIMARY)
        self._disabled_episodes.clear()
        self._xlsx_check.config(state="normal", fg=TEXT_PRIMARY)

        # Restore full export list from engine
        try:
            self._suppress_progress = True
            full_list = self.engine.prepare_export(
                self.show_code.get().strip() or "SHOW",
                self.show_acronym.get().strip() or "ACRN",
                self.export_date.get().strip() or datetime.now().strftime("%y%m%d")
            )
            self._suppress_progress = False
            self.export_list = full_list
            self.engine.export_list = full_list
        except Exception:
            self._suppress_progress = False

        self._update_episode_list()
        self._update_plate_count()
        self.btn_export._text = "Export"
        self._set_progress_normal()
        self.progress_var.set(0)
        self.progress_status.config(text="Click Export to Continue.")
        self._set_btn_state(self.btn_export, True)
        self._log("Export reset. Ready to export from the beginning.", "muted")
        if hasattr(self, 'step_canvases') and len(self.step_canvases) > 2:
            c = self.step_canvases[2]
            size = 24
            c.delete("all")
            c.create_oval(1, 1, size-1, size-1, outline=ACCENT, width=2, fill="")
            c.create_text(size//2, size//2, text="3",
                         font=("SF Pro Display", 12, "bold"), fill=ACCENT)

    def _show_continue_buttons(self, stopped_at):
        """Change Export button to Continue, enable Reset Export."""
        def _do():
            self._enable_reset_export_btn()
            self.btn_export._text = "Continue"
            self.btn_export._bg = ACCENT
            self.btn_export._fg = "#000000"
            self.btn_export._draw(ACCENT, "#000000")
            self.btn_export.unbind("<ButtonPress-1>")
            self.btn_export.unbind("<ButtonRelease-1>")
            self.btn_export.unbind("<Enter>")
            self.btn_export.unbind("<Leave>")
            def _on_press(e):
                self.btn_export._draw("#c07010", "#000000")
            def _on_release(e):
                self.btn_export._draw(ACCENT, "#000000")
                self._set_btn_state(self.btn_export, False)
                self._do_continue_export()
            self.btn_export.bind("<ButtonPress-1>", _on_press)
            self.btn_export.bind("<ButtonRelease-1>", _on_release)
            self.btn_export.bind("<Enter>", lambda e: self.btn_export._draw(ACCENT_HOVER, "#000000"))
            self.btn_export.bind("<Leave>", lambda e: self.btn_export._draw(ACCENT, "#000000"))
            # Show Reset Export - already always visible
        self.after(0, _do)

    def _hide_continue_buttons(self):
        def _do():
            # Restore Export button
            self.btn_export._text = "Export"
            self._set_btn_state(self.btn_export, False)
            self.btn_export.unbind("<Button-1>")
        self.after(0, _do)

    def _set_stop_button(self):
        """Set RESET ALL button to STOP state for export."""
        self.btn_reset._text = "STOP"
        self.btn_reset._bg = "#8e2a2a"
        self.btn_reset._draw("#8e2a2a")
        # Clear all existing bindings first
        self.btn_reset.unbind("<Enter>")
        self.btn_reset.unbind("<Leave>")
        self.btn_reset.unbind("<ButtonPress-1>")
        self.btn_reset.unbind("<ButtonRelease-1>")
        self.btn_reset.bind("<Enter>", lambda e: self.btn_reset._draw("#ae3a3a"))
        self.btn_reset.bind("<Leave>", lambda e: self.btn_reset._draw(self.btn_reset._bg))
        self.btn_reset.bind("<ButtonPress-1>", lambda e: self.btn_reset._draw("#6e1a1a"))
        def _on_stop_release(e):
            self.btn_reset._text = "RESET ALL"
            self.btn_reset._bg = "#2a6e2a"
            self.btn_reset._draw("#2a6e2a")
            self._enable_reset_btn()
            self._stop_export_now()
        self.btn_reset.bind("<ButtonRelease-1>", _on_stop_release)

    def _run_export_task(self, output, start_index=0):
        """Run export in a background thread."""
        self._stop_export = False
        self._set_progress_normal()
        if start_index == 0:
            self.progress_var.set(0)
            self.progress_status.config(text="")
        self._set_btn_state(self.btn_export, False)
        self._set_stop_button()
        self._start_thinking()
        self._export_started = True
        self.after(0, self._update_episode_list)
        self.after(0, lambda: self.btn_toggle_all.config(bg="#2a2a2a", fg="#555555"))
        self.after(0, lambda: self._xlsx_check.config(state="disabled", fg="#444444"))

        def task():
            try:
                # Always sync filtered list to engine before running
                self.engine.export_list = self.export_list

                # Setup per-clip screenshots if checkbox enabled
                _do_screenshots = self._create_xlsx.get()
                _shot_map = {}
                _list_folder = os.path.join(output, "LIST TO POST")
                _shots_folder = os.path.join(_list_folder, "SCREENSHOTS")
                _grab_screenshot = None

                if _do_screenshots:
                    import cv2 as _cv2
                    from PIL import Image as _PILImage
                    _ref = self.reference_video_path.get().strip() or None
                    _cap = _cv2.VideoCapture(_ref) if _ref and os.path.exists(str(_ref)) else None
                    if _cap and not _cap.isOpened():
                        _cap = None
                        self._log("  ⚠ Could not open ref video for screenshots.", "warn")
                    else:
                        os.makedirs(_shots_folder, exist_ok=True)
                        _start_tc = self.start_timecode.get().strip() or "03:59:50:00"
                        _fps = float(self.engine.timeline.GetSetting("timelineFrameRate") or 24)
                        _video_start_frames = timecode_str_to_frames(_start_tc, _fps)
                        # ref video opened silently — will log per clip during export

                        def _grab_screenshot(item):
                            try:
                                file_frame = int(item["start_frame"] - _video_start_frames)
                                if file_frame < 0:
                                    return
                                _cap.set(_cv2.CAP_PROP_POS_FRAMES, file_frame)
                                ret, frame = _cap.read()
                                if not ret or frame is None:
                                    self._log(f"  ⚠ Screenshot: could not read frame {file_frame}", "warn")
                                    return
                                rgb = _cv2.cvtColor(frame, _cv2.COLOR_BGR2RGB)
                                pil = _PILImage.fromarray(rgb)
                                # Save at 720x1280 (9:16 vertical 720p)
                                pil = pil.resize((720, 1280), _PILImage.LANCZOS)
                                shot_path = os.path.join(_shots_folder, item["filename"] + ".jpg")
                                pil.save(shot_path, "JPEG", quality=95)
                                _shot_map[item["filename"]] = (shot_path, 720, 1280)
                                self._log(f"  📷 Screenshot saved: {item['filename']} (frame {file_frame})", "muted")
                            except Exception as e:
                                self._log(f"  ⚠ Screenshot error: {e}", "warn")

                success = self.engine.run_export(output,
                    stop_flag=lambda: self._stop_export,
                    start_index=start_index,
                    screenshot_callback=_grab_screenshot)

                if _do_screenshots and _cap:
                    _cap.release()

                self.after(0, self._stop_thinking)
                def _to_reset():
                    self.btn_reset._text = "RESET ALL"
                    self.btn_reset._bg = "#2a6e2a"
                    self.btn_reset._draw("#2a6e2a")
                    self._enable_reset_btn()
                self.after(0, _to_reset)

                if self._stop_export:
                    done = self.engine._last_done if hasattr(self.engine, '_last_done') else start_index
                    total_plates = len(self.export_list)
                    last_item = self.export_list[done - 1] if done > 0 else {}
                    last_ep = last_item.get("episode_code", "?")
                    last_track = last_item.get("track_index", "?")
                    self._log(f"⚠ Stopped at plate {done}/{total_plates} · {last_ep} · Track {last_track}", "warn")
                    self.after(0, lambda d=done, t=total_plates, ep=last_ep, tr=last_track: self.progress_status.config(
                        text=f"Stopped at plate {d}/{t} · {ep} · Track {tr}"))
                    self.after(0, lambda: self._enable_reset_export_btn())
                    self._show_continue_buttons(done)
                elif success:
                    self._log("✓ Export complete.", "success")
                    self._log(f"  Output: {output}", "muted")
                    self._update_progress(100, None)
                    self._set_progress_green()
                    self._set_step_done(2)
                    self.after(0, lambda: self.progress_status.config(
                        text="All VFX plates have been exported."))
                    self.after(0, self._enable_reset_export_btn)
                    if _do_screenshots and _shot_map:
                        self.after(0, lambda: self._log("Building plate list xlsx...", "muted"))
                        _exp_list = list(self.export_list)
                        _show = self.show_code.get().strip() or "SHOW"
                        _acr = self.show_acronym.get().strip() or "ACRN"
                        _date = self.export_date.get().strip() or datetime.now().strftime("%y%m%d")
                        _tl = self.engine.timeline
                        _lf = _list_folder
                        _sm = dict(_shot_map)
                        def _safe_log(msg, tag="muted"):
                            self.after(0, lambda m=msg, t=tag: self._log(m, t))
                        def _gen_xlsx():
                            try:
                                generate_plate_list_xlsx(
                                    export_list=_exp_list,
                                    output_dir=output,
                                    show_code=_show,
                                    acronym=_acr,
                                    date_str=_date,
                                    shot_map=_sm,
                                    list_folder=_lf,
                                    timeline=_tl,
                                    log_callback=_safe_log
                                )
                            except Exception as xe:
                                _safe_log(f"✗ Plate list error: {xe}", "error")
                        threading.Thread(target=_gen_xlsx, daemon=True).start()
                else:
                    self._log("✗ Export failed — check log for details.", "error")
                    self.after(0, lambda: self.progress_status.config(
                        text="Export failed — check log for details."))
                    def _show_reset_only():
                        self._set_btn_state(self.btn_export, False)
                        self._enable_reset_export_btn()
                    self.after(0, _show_reset_only)
            except Exception as e:
                print(f"TASK EXCEPTION: {e}", flush=True)
                import traceback; traceback.print_exc()
                self.after(0, lambda err=str(e): self._log(f"✗ {err}", "error"))
                self.after(0, self._stop_thinking)
                self.after(0, lambda: self._set_btn_state(self.btn_export, True))
        threading.Thread(target=task, daemon=True).start()

    def _switch_tab(self, tab_id):
        self._active_tab = tab_id
        if hasattr(self, '_draw_tab_fn'):
            self._draw_tab_fn()
        if tab_id == "vfx":
            self._test_content.pack_forget()
            self._vfx_content.pack(fill="both", expand=True)
        elif tab_id == "test":
            self._vfx_content.pack_forget()
            self._test_content.pack(fill="both", expand=True)

    def _toggle_log(self):
        if self._log_visible:
            self.log_frame_outer.pack_forget()
            self.btn_log_toggle.config(text="SHOW LOG")
            self._log_visible = False
        else:
            self.log_frame_outer.pack(fill="x")
            self.btn_log_toggle.config(text="HIDE LOG")
            self._log_visible = True

    def _set_step_done(self, step_index):
        def _do():
            if step_index < len(self.step_canvases):
                c = self.step_canvases[step_index]
                size = 24
                c.delete("all")
                c.create_oval(1, 1, size-1, size-1, outline=TEXT_SUCCESS, width=2, fill="")
                c.create_text(size//2, size//2, text="✓",
                             font=("SF Pro Display", 13, "bold"), fill=TEXT_SUCCESS)
        self.after(0, _do)

    def _set_step_active(self, step_index):
        def _do():
            if step_index < len(self.step_canvases):
                c = self.step_canvases[step_index]
                size = 24
                c.delete("all")
                c.create_oval(1, 1, size-1, size-1, outline=ACCENT, width=2, fill=ACCENT)
                c.create_text(size//2, size//2, text=str(step_index + 1),
                             font=("SF Pro Display", 12, "bold"), fill=BG_DARK)
        self.after(0, _do)

    def _set_progress_green(self):
        def _do():
            s = ttk.Style()
            s.configure("TProgressbar", background=TEXT_SUCCESS)
        self.after(0, _do)

    def _set_progress_normal(self):
        def _do():
            s = ttk.Style()
            s.configure("TProgressbar", background=ACCENT)
        self.after(0, _do)

    def _mode_row(self, parent, var, command):
        row = tk.Frame(parent, bg=BG_PANEL)
        row.pack(fill="x", pady=(0, 6))
        tk.Label(row, text="Input Mode", font=FONT_LABEL, bg=BG_PANEL,
                 fg=TEXT_PRIMARY, width=22, anchor="w").pack(side="left")
        for mode, label in [("auto", "Auto"), ("manual", "Manual")]:
            rb = tk.Radiobutton(row, text=label, variable=var, value=mode,
                                command=command, font=FONT_SMALL, bg=BG_PANEL,
                                fg=TEXT_PRIMARY, selectcolor=BG_INPUT,
                                activebackground=BG_PANEL, activeforeground=TEXT_PRIMARY,
                                indicatoron=True, relief="flat", bd=0)
            rb.pack(side="left", padx=(0, 20))

    def _set_widgets_enabled(self, widgets, enabled):
        for w in widgets:
            if isinstance(w, tk.Entry):
                w.config(state="normal" if enabled else "disabled",
                         fg=TEXT_PRIMARY if enabled else TEXT_MUTED,
                         bg=BG_INPUT if enabled else "#2a2a2a")
            elif hasattr(w, "_enabled"):
                self._set_btn_state(w, enabled)

    def _on_show_mode_change(self):
        if hasattr(self, "_enable_reset_btn"): self._enable_reset_btn()
        is_auto = self.show_mode.get() == "auto"
        if is_auto:
            self.show_manual_frame.pack_forget()
            self.hour_wrap.pack_forget()
            self.hour_hint.pack_forget()
            self.hour_display.pack(side="left")
        else:
            self.show_manual_frame.pack(fill="x")
            self.hour_display.pack_forget()
            self.hour_wrap.pack(side="left")
            self.hour_hint.pack(side="left", padx=(8, 0))
        self._update_filename_preview()

    def _on_ref_mode_change(self):
        if hasattr(self, "_enable_reset_btn"): self._enable_reset_btn()
        is_auto = self.ref_mode.get() == "auto"
        self._set_widgets_enabled([self.ref_entry, self.ref_browse_btn], not is_auto)
        self.ref_hint.config(text="Auto-detected on connect" if is_auto else "Browse to select your reference video")

    def _on_out_mode_change(self):
        if hasattr(self, "_enable_reset_btn"): self._enable_reset_btn()
        is_auto = self.out_mode.get() == "auto"
        self._set_widgets_enabled([self.out_entry, self.out_browse_btn], not is_auto)
        self.out_hint.config(text="Auto-detected on connect" if is_auto else "Browse to select your output folder")

    def _rounded_btn(self, parent, text, command, enabled=True, accent=False, small=False, reserve_text=None, match_height=False):
        """Button drawn on Canvas with smooth rounded corners."""
        if accent and not enabled:
            bg, fg = "#5a4a1a", "#888888"
        elif accent:
            bg, fg = ACCENT, "#000000"
        else:
            bg, fg = ("#444444" if enabled else "#333333"), ("#FFFFFF" if enabled else "#555555")

        font = ("SF Pro Display", 11) if small else ("SF Pro Display", 13)
        pad_x, pad_y = (12, 4) if small else (18, 7)
        r = 8
        if match_height:
            # Will be resized after packing via place or configure
            pad_y = 7  # match entry ipady=7

        # Size canvas using font metrics - no temporary widget needed
        size_text = reserve_text if reserve_text else text
        import tkinter.font as tkfont
        f = tkfont.Font(family="SF Pro Display", size=11 if small else 13)
        tw = f.measure(size_text)
        th = f.metrics("linespace")
        w = tw + pad_x * 2
        h = th + pad_y * 2

        c = tk.Canvas(parent, width=w, height=h, bg=parent.cget("bg"),
                      highlightthickness=0, cursor="" if enabled else "arrow")

        # Initialize _text before first draw
        c._text = text

        # Draw once with tags for fast itemconfig updates
        x1, y1, x2, y2 = 0, 0, w, h
        points = [
            x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
            x2, y2-r, x2, y2, x2-r, y2, x1+r, y2,
            x1, y2, x1, y2-r, x1, y1+r, x1, y1,
        ]
        _poly_id = c.create_polygon(points, fill=bg, outline="", smooth=True)
        _text_id = c.create_text(w//2, h//2, text=c._text, font=font, fill=fg)

        def _draw(fill, text_color):
            c.itemconfig(_poly_id, fill=fill)
            c.itemconfig(_text_id, fill=text_color, text=c._text)

        # Store mutable state
        state = {"bg": bg, "fg": fg}

        if enabled:
            hover = ACCENT_HOVER if accent else "#555555"
            press = "#c07010" if accent else "#666666"
            c.bind("<Enter>", lambda e: _draw(hover, state["fg"]))
            c.bind("<Leave>", lambda e: _draw(state["bg"], state["fg"]))
            c.bind("<ButtonPress-1>", lambda e: _draw(press, state["fg"]))
            def _on_btn_release(e, cmd=command, s=state):
                _draw(s["bg"], s["fg"])
                cmd()
            c.bind("<ButtonRelease-1>", _on_btn_release)

        c._bg = bg
        c._fg = fg
        c._accent = accent
        c._enabled = enabled
        c._command = command
        c._state = state

        def _redraw(fill_color, text_color):
            state["bg"] = fill_color
            state["fg"] = text_color
            _draw(fill_color, text_color)

        c._draw = _redraw
        c._text = text
        c._font = font
        return c

    def _section_label(self, parent, text):
        tk.Label(parent, text=text, font=("SF Pro Display", 10, "bold"),
                 bg=BG_DARK, fg=TEXT_MUTED).pack(anchor="w", pady=(12, 4))

    def _panel(self, parent):
        f = tk.Frame(parent, bg=BG_PANEL, bd=0, highlightthickness=0)
        f.pack(fill="x", pady=(0, 4))
        inner = tk.Frame(f, bg=BG_PANEL, padx=12, pady=10)
        inner.pack(fill="both", expand=True)
        return inner

    def _field_row(self, parent, label, var, placeholder=""):
        row = tk.Frame(parent, bg=BG_PANEL)
        row.pack(fill="x", pady=4)
        tk.Label(row, text=label, font=FONT_LABEL, bg=BG_PANEL,
                 fg=TEXT_PRIMARY, width=22, anchor="w").pack(side="left")
        # Entry wrapper with left padding via bg-matched frame
        entry_wrap = tk.Frame(row, bg=BG_INPUT,
                              highlightthickness=1,
                              highlightbackground=BORDER,
                              highlightcolor=ACCENT)
        entry_wrap.pack(side="left", fill="x", expand=True)
        tk.Frame(entry_wrap, bg=BG_INPUT, width=2).pack(side="left")
        entry = tk.Entry(entry_wrap, font=FONT_SMALL,
                         bg=BG_INPUT, fg=TEXT_PRIMARY, insertbackground=TEXT_PRIMARY,
                         relief="flat", bd=0, highlightthickness=0)
        entry.pack(side="left", fill="x", expand=True, ipady=7)

        def sync_var_to_entry(*args):
            current = entry.get()
            val = var.get()
            if current != val:
                entry.delete(0, "end")
                if val:
                    entry.insert(0, val)
                    entry.config(fg=TEXT_PRIMARY)
                elif placeholder:
                    entry.insert(0, placeholder)
                    entry.config(fg=TEXT_MUTED)
        var.trace_add("write", sync_var_to_entry)

        if var.get():
            entry.insert(0, var.get())
        elif placeholder:
            entry.insert(0, placeholder)
            entry.config(fg=TEXT_MUTED)

        def on_in(e, ent=entry, ph=placeholder):
            if ent.get() == ph and ent.cget("fg") == TEXT_MUTED:
                ent.delete(0, "end")
                ent.config(fg=TEXT_PRIMARY)

        def on_out(e, ent=entry, ph=placeholder, v=var):
            val = ent.get().strip()
            if not val:
                if v.get():
                    ent.insert(0, v.get())
                    ent.config(fg=TEXT_PRIMARY)
                elif ph:
                    ent.insert(0, ph)
                    ent.config(fg=TEXT_MUTED)
            else:
                v.set(val)
                ent.config(fg=TEXT_PRIMARY)

        def on_key(e, ent=entry, v=var):
            v.set(ent.get())

        entry.bind("<FocusIn>", on_in)
        entry.bind("<FocusOut>", on_out)
        entry.bind("<KeyRelease>", on_key)

    def _btn(self, parent, text, command, enabled=True, accent=False, small=False):
        if accent and not enabled:
            bg = "#5a4a1a"  # dark muted orange when disabled
            fg = "#888888"
        elif accent:
            bg = ACCENT
            fg = "#000000"
        else:
            bg = "#444444" if enabled else "#333333"
            fg = "#FFFFFF" if enabled else "#555555"

        btn = tk.Label(parent, text=text,
                       font=("SF Pro Display", 11) if small else ("SF Pro Display", 13),
                       bg=bg, fg=fg,
                       padx=12 if small else 18, pady=5 if small else 9,
                       cursor="" if enabled else "arrow")
        if enabled:
            def on_enter(e, b=btn, a=accent): b.config(bg=ACCENT_HOVER if a else "#555555")
            def on_leave(e, b=btn, ob=bg): b.config(bg=ob)
            def on_click(e, cmd=command): cmd()
            btn.bind("<Enter>", on_enter)
            btn.bind("<Leave>", on_leave)
            btn.bind("<Button-1>", on_click)
        btn._enabled = enabled
        btn._bg = bg
        btn._accent = accent
        btn._command = command
        return btn

    def _on_naming_mode_change(self):
        is_auto = self.naming_mode.get() == "auto"
        if is_auto:
            self.manual_frame.pack_forget()
            self.hour_wrap.pack_forget()
            self.hour_hint.pack_forget()
            self.hour_display.pack(side="left")
            # Disable ref video section in auto mode
            self.ref_video_frame.config(
                highlightbackground="#222222")
            for w in self.ref_video_widgets:
                try: w.config(state="disabled")
                except: pass
        else:
            self.manual_frame.pack(fill="x")
            self.hour_display.pack_forget()
            self.hour_wrap.pack(side="left")
            self.hour_hint.pack(side="left", padx=(8, 0))
            # Enable ref video section in manual mode
            self.ref_video_frame.config(
                highlightbackground=BORDER)
            for w in self.ref_video_widgets:
                try: w.config(state="normal")
                except: pass
        self._update_filename_preview()

    def _update_filename_preview(self):
        code = self.show_code.get().strip()
        acronym = self.show_acronym.get().strip()
        date = self.export_date.get().strip()
        if code and acronym and date:
            self.filename_preview.set(f"{code}_{acronym}_VFX_EP##_{date}_##")
        else:
            self.filename_preview.set("Connect to DaVinci to auto-fill...")

    def _check_deps_on_start(self):
        missing = check_dependencies()
        if missing:
            self._log("Missing dependencies:", "warn")
            for d in missing:
                self._log(f"  • {d}", "warn")
        else:
            self._log("All dependencies OK. Ready to connect.", "success")

    def _validate_show_fields(self):
        if not self.show_code.get().strip():
            messagebox.showerror("Missing Info", "Please enter a Show Code.")
            return False
        if not self.show_acronym.get().strip():
            messagebox.showerror("Missing Info", "Please enter a Show Acronym.")
            return False
        if not self.export_date.get().strip():
            messagebox.showerror("Missing Info", "Please enter an Export Date.")
            return False
        return True

    def _stop_scan_now(self):
        self._stop_scan = True
        self._log("⚠ Stopping scan after current frame...", "warn")

    def _start_thinking(self):
        """Start GIF animation."""
        if self._thinking_active:
            return
        self._thinking_active = True
        self._thinking_frame_idx = 0
        if self._thinking_frames:
            self._animate_gif()
        else:
            # Fallback: simple text dots animation
            self._animate_text_dots(0)

    def _stop_thinking(self):
        """Stop GIF animation and hide."""
        self._thinking_active = False
        if hasattr(self, '_thinking_label'):
            self.after(0, lambda: self._thinking_label.config(image="", text=""))

    def _animate_gif(self):
        if not self._thinking_active or not self._thinking_frames:
            return
        frame = self._thinking_frames[self._thinking_frame_idx]
        self._thinking_label.config(image=frame)
        self._thinking_frame_idx = (self._thinking_frame_idx + 1) % len(self._thinking_frames)
        self.after(80, self._animate_gif)

    def _animate_text_dots(self, frame):
        if not self._thinking_active:
            self._thinking_label.config(text="", image="")
            return
        dots = ["●  ○  ○", "○  ●  ○", "○  ○  ●"]
        self._thinking_label.config(
            image="",
            text=dots[frame % 3],
            font=("SF Pro Display", 9),
            fg="#FFFFFF",
            bg=BG_OUTER
        )
        self.after(250, lambda f=frame: self._animate_text_dots(f + 1))

    def _check_for_updates(self):
        """Silently check GitHub for a newer version."""
        def task():
            try:
                import urllib.request, json, ssl
                ctx = ssl.create_default_context()
                ctx.check_hostname = False
                ctx.verify_mode = ssl.CERT_NONE
                req = urllib.request.urlopen(VERSION_URL, context=ctx, timeout=10)
                raw = req.read().decode()
                with open(os.path.expanduser("~/Desktop/ft_update_log.txt"), "a") as lf:
                    lf.write(f"Response: {raw}\n")
                data = json.loads(raw)
                remote = data.get("version", "0")
                notes  = data.get("release_notes", "")
                with open(os.path.expanduser("~/Desktop/ft_update_log.txt"), "a") as lf:
                    lf.write(f"Remote: {remote}, Local: {APP_VERSION}\n")
                def _parse(v):
                    try: return tuple(int(x) for x in v.split("."))
                    except: return (0,)
                if _parse(remote) > _parse(APP_VERSION):
                    self.after(0, lambda: self._show_update_banner(remote, notes, data.get("download_url", DOWNLOAD_URL)))
                else:
                    with open(os.path.expanduser("~/Desktop/ft_update_log.txt"), "a") as lf:
                        lf.write(f"No update needed: {remote} <= {APP_VERSION}\n")
            except Exception as e:
                with open(os.path.expanduser("~/Desktop/ft_update_log.txt"), "a") as lf:
                    lf.write(f"Update check failed: {e}\n")
        threading.Thread(target=task, daemon=True).start()

    def _show_update_banner(self, remote_version, notes, download_url):
        """Show a dismissable update banner below the title."""
        BNR_BG  = "#1a3a1a"
        BNR_FG  = "#6fcf6f"
        BTN_BG  = "#2a6e2a"
        BTN_HOV = "#3a8e3a"
        BTN_PRS = "#1a5e1a"

        banner = tk.Frame(self._main_frame, bg=BNR_BG)
        banner.pack(fill="x", pady=(0, 8))

        def _make_banner_btn(parent, text, cmd):
            b = tk.Label(parent, text=text, font=("SF Pro Display", 11),
                         bg=BTN_BG, fg="#ffffff", padx=10, pady=4, cursor="")
            b.bind("<Enter>",           lambda e: b.config(bg=BTN_HOV))
            b.bind("<Leave>",           lambda e: b.config(bg=BTN_BG))
            b.bind("<ButtonPress-1>",   lambda e: b.config(bg=BTN_PRS))
            b.bind("<ButtonRelease-1>", lambda e: (b.config(bg=BTN_HOV), cmd()))
            return b

        # ✕ on the left
        _make_banner_btn(banner, "✕", banner.destroy).pack(side="left", padx=(8, 4), pady=4)

        # Message text
        tk.Label(banner, text=f"✦ Version {remote_version} available — {notes}",
                 font=("SF Pro Display", 11), bg=BNR_BG, fg=BNR_FG, padx=8, pady=6
                 ).pack(side="left")

        # UPDATE NOW on the right
        _make_banner_btn(banner, "UPDATE NOW", lambda: self._do_update(remote_version, download_url)
                         ).pack(side="right", padx=(4, 8), pady=4)

    def _do_update(self, remote_version, download_url):
        import urllib.request, ssl
        self._log(f"Downloading v{remote_version}...", "muted")
        try:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE

            script_path = os.path.abspath(__file__)
            resources_dir = os.path.dirname(script_path)
            vpath = os.path.join(resources_dir, "version.json")

            # Download version.json — same path resolution as main.py
            vreq = urllib.request.urlopen(VERSION_URL, context=ctx, timeout=30)
            with open(vpath, "wb") as f:
                f.write(vreq.read())
            self._log(f"  version.json saved ✓", "muted")

            # Download main.py
            req = urllib.request.urlopen(download_url, context=ctx, timeout=30)
            with open(script_path, "wb") as f:
                f.write(req.read())
            self._log(f"  main.py saved ✓", "muted")
            self._log(f"✓ Updated to v{remote_version}. Restarting...", "success")
            # Find the .app bundle by walking up from script_path
            path = os.path.abspath(script_path)
            app_bundle = None
            for _ in range(6):
                path = os.path.dirname(path)
                if path.endswith(".app"):
                    app_bundle = path
                    break
            def _restart():
                if app_bundle and os.path.exists(app_bundle):
                    # Use the app's own launcher executable directly
                    # -n forces a new instance even if already running
                    import subprocess
                    app_name = os.path.splitext(os.path.basename(app_bundle))[0]
                    launcher = os.path.join(app_bundle, "Contents", "MacOS", app_name)
                    if os.path.exists(launcher):
                        subprocess.Popen([launcher])
                    else:
                        os.system(f'open -n "{app_bundle}"')
                else:
                    os.execv(sys.executable, [sys.executable, script_path])
                self.after(500, self.destroy)
            self.after(1500, _restart)
        except Exception as e:
            self._log(f"✗ Update failed: {e}", "error")

    def _enable_reset_btn(self):
        """Enable the RESET ALL button."""
        if not hasattr(self, 'btn_reset'):
            return
        self.btn_reset._text = "RESET ALL"
        self.btn_reset._bg = "#2a6e2a"
        self.btn_reset._draw("#2a6e2a")
        self.btn_reset._action = self._do_reset
        self.btn_reset.bind("<Enter>", lambda e: self.btn_reset._draw("#3a8e3a"))
        self.btn_reset.bind("<Leave>", lambda e: self.btn_reset._draw(self.btn_reset._bg))
        self.btn_reset.bind("<ButtonPress-1>", lambda e: self.btn_reset._draw("#1a5e1a"))
        def _on_release(e):
            if self.btn_reset._text == "STOP":
                self.btn_reset._text = "RESET ALL"
                self.btn_reset._bg = "#2a6e2a"
                self.btn_reset._draw("#2a6e2a")
                self.btn_reset.after(10, self.btn_reset._action)
            else:
                self._disable_reset_btn()
                self.btn_reset.after(10, self.btn_reset._action)
        self.btn_reset.bind("<ButtonRelease-1>", _on_release)

    def _disable_reset_btn(self):
        """Disable the RESET ALL button - greyed out."""
        if not hasattr(self, 'btn_reset'):
            return
        self.btn_reset._draw("#2a2a2a")
        self.btn_reset.unbind("<Enter>")
        self.btn_reset.unbind("<Leave>")
        self.btn_reset.unbind("<ButtonPress-1>")
        self.btn_reset.unbind("<ButtonRelease-1>")

    def _restore_reset_btn(self):
        def _do():
            self.btn_reset._text = "RESET ALL"
            self.btn_reset._bg = "#2a6e2a"
            self.btn_reset._draw("#2a6e2a")
            self._enable_reset_btn()
        self.after(0, _do)

    def _stop_export_now(self):
        self._stop_export = True
        self._log("⚠ Stop requested — will stop after current clip.", "warn")
        self.after(0, lambda: self.progress_status.config(
            text="Stop requested — finishing current clip..."))

    def _do_reset(self):
        """Full reset — clear everything and go back to step 1."""
        self.after(0, self._disable_reset_btn)
        self._stop_export = True
        self._stop_scan = True
        self._export_started = False
        self._all_disabled = False
        self.btn_toggle_all.config(text="DISABLE ALL", bg=BG_INPUT, fg=TEXT_PRIMARY)
        self._disabled_episodes.clear()
        self._create_xlsx.set(True)
        self._xlsx_check.config(state="normal", fg=TEXT_PRIMARY)
        self.engine = None
        self.episode_markers = []
        self.export_list = []

        # Reset step circles
        for i, c in enumerate(self.step_canvases):
            size = 24
            c.delete("all")
            c.create_oval(1, 1, size-1, size-1, outline=ACCENT, width=2, fill="")
            c.create_text(size//2, size//2, text=str(i+1),
                         font=("SF Pro Display", 12, "bold"), fill=ACCENT)

        # Reset progress bar and status
        self._set_progress_normal()
        self.progress_var.set(0)
        self.progress_status.config(text="")

        # Reset text vars
        self.filename_preview.set("Connect to DaVinci to auto-fill...")
        self.start_timecode_display.set("Auto-detected on connect")
        self.hour_display.config(fg=TEXT_MUTED)

        # Clear log
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")

        # Clear episode tags
        for w in self._ep_tag_widgets:
            try: w.destroy()
            except: pass
        self._ep_tag_widgets.clear()

        # Reset plates count
        self.plates_count_label.config(text="—")

        # Disable Reset Export button
        self._set_btn_state(self.btn_reset_export, False)

        # Restore export button text
        self.btn_export._text = "Export"

        # Reset all step buttons
        self._set_btn_state(self.btn_connect, True)
        self._set_btn_state(self.btn_scan, False)
        self._set_btn_state(self.btn_export, False)

        # Reset RESET ALL button - stays disabled after full reset
        self.btn_reset._text = "RESET ALL"
        self.btn_reset._bg = "#2a6e2a"
        self.btn_reset._draw("#2a2a2a")  # show greyed out

        self._log("Reset. Ready to connect.", "success")

    def _do_connect(self):
        self._enable_reset_btn()
        self.after(0, lambda: self._connect_status.config(text="Connecting..."))
        def task():
            try:
                self._set_step_active(0)
                self.engine = ExportEngine(log_callback=self._log,
                                           progress_callback=self._update_progress)
                info = self.engine.connect()
                self._set_btn_state(self.btn_connect, False)
                self._log(f"✓ Connected: {info['name']}.", "success")
                self._set_step_done(0)
                self.after(0, lambda: self._connect_status.config(text=""))

                def autofill():
                    # SHOW INFO section
                    if self.show_mode.get() == "auto":
                        self.show_code.set(info.get("show_code", ""))
                        self.show_acronym.set(info.get("show_acronym", ""))
                        self.export_date.set(datetime.now().strftime("%y%m%d"))
                        full_tc = info.get("start_timecode", "03:59:50:00")
                        self.start_timecode.set(full_tc)
                        self.start_timecode_display.set(full_tc)
                        self.hour_display.config(fg=ACCENT)
                        self._update_filename_preview()

                    # REFERENCE VIDEO section
                    if self.ref_mode.get() == "auto":
                        ref_path = get_offline_reference_path(
                            self.engine.project,
                            show_code=self.show_code.get().strip()
                        )
                        if ref_path:
                            self.reference_video_path.set(ref_path)
                            self._log(f"Ref video: {os.path.basename(ref_path)}.", "success")
                        else:
                            self._log("Reference video not found — will try again on scan.", "warn")

                    # Enable scan only after all detection is complete
                    self.after(0, lambda: self._set_btn_state(self.btn_scan, True))
                    self.after(0, self._stop_thinking)
                    self.after(0, lambda: self._xlsx_check.config(state="normal", fg=TEXT_PRIMARY))

                    # OUTPUT FOLDER section
                    if self.out_mode.get() == "auto":
                        found = find_output_folder(
                            info.get("show_code", ""),
                            info.get("show_acronym", "")
                        )
                        if found:
                            self.output_dir.set(found)
                            self._log(f"Output folder: {found}", "success")
                        else:
                            self._log("Output folder not found — switching to Manual.", "warn")
                            self.out_mode.set("manual")
                            self._on_out_mode_change()
                self.after(0, autofill)

            except Exception as e:
                self._log(f"✗ {e}", "error")
                self.after(0, lambda: self._connect_status.config(text=""))
        threading.Thread(target=task, daemon=True).start()

    def _do_scan(self):
        # Immediate UI feedback
        self._xlsx_check.config(state="disabled", fg="#444444")
        self._set_btn_state(self.btn_scan, False)
        self.progress_status.config(text="")
        self.progress_var.set(0)
        self._set_progress_normal()
        self._set_step_active(1)
        self._start_thinking()
        self.btn_reset._text = "STOP"
        self.btn_reset._bg = "#8e2a2a"
        self.btn_reset._draw("#8e2a2a")
        self.btn_reset._action = self._stop_scan_now
        self.btn_reset.bind("<Enter>", lambda e: self.btn_reset._draw("#ae3a3a"))
        self.btn_reset.bind("<Leave>", lambda e: self.btn_reset._draw(self.btn_reset._bg))
        self.btn_reset.bind("<ButtonPress-1>", lambda e: self.btn_reset._draw("#6e1a1a"))
        def _stop_scan_release(e):
            self.btn_reset._text = "RESET ALL"
            self.btn_reset._bg = "#2a6e2a"
            self.btn_reset._draw("#2a6e2a")
            self.btn_reset._action = self._do_reset
            self._enable_reset_btn()
            self._stop_scan_now()
        self.btn_reset.bind("<ButtonRelease-1>", _stop_scan_release)
        def task():
            try:
                self._scanning = True

                # Pass ref video path regardless of mode - already detected during connect
                ref = self.reference_video_path.get().strip() or None
                tc = self.start_timecode.get().strip() or "03:59:50:00"
                self._stop_scan = False
                self.episode_markers = self.engine.scan_episodes(
                    ref, start_timecode=tc, stop_flag=lambda: self._stop_scan)
                self._scanning = False
                self.after(0, self._stop_thinking)

                if self._stop_scan:
                    self._scanning = False
                    self._set_progress_normal()
                    self.after(0, lambda: self.progress_var.set(0))
                    self._log(f"⚠ Scan stopped. Found {len(self.episode_markers)} episodes so far.", "warn")
                    self.after(0, lambda: self.progress_status.config(
                        text="Scan stopped. Click Scan Episodes to retry."))
                    self.after(0, lambda: self._xlsx_check.config(state="normal", fg=TEXT_PRIMARY))
                    self._set_btn_state(self.btn_scan, True)
                    return

                # Scan completed successfully
                self._log(f"✓ Detected {len(self.episode_markers)} episodes", "success")

                # Silently count plates (suppress progress to avoid bar dipping)
                try:
                    self._suppress_progress = True
                    code = self.show_code.get().strip()
                    acronym = self.show_acronym.get().strip()
                    date = self.export_date.get().strip()
                    if not code: code = "SHOW"
                    if not acronym: acronym = "ACRN"
                    if not date: date = datetime.now().strftime("%y%m%d")
                    preview = self.engine.prepare_export(code, acronym, date)
                    self._suppress_progress = False
                    self.export_list = preview
                    self.after(0, self._update_plate_count)
                    self._log(f"  {len(preview)} plates ready to export.", "muted")
                except Exception as e:
                    self._suppress_progress = False
                    self._log(f"  Could not count plates: {e}", "muted")

                # Hit 100% first, then show episodes
                self._update_progress(100, None)
                self._set_progress_green()
                self._set_step_done(1)
                self._update_episode_list()
                self.after(0, lambda: self.progress_status.config(text="Click Export to Continue."))
                self.after(0, self._restore_reset_btn)
                self.after(0, lambda: self._set_btn_state(self.btn_scan, False))
                self._set_btn_state(self.btn_export, True)
                self.after(0, lambda: self._xlsx_check.config(state="normal", fg=TEXT_PRIMARY))
            except FileNotFoundError:
                self._scanning = False
                self.after(0, self._stop_thinking)
                self._log("✗ Reference video not found. Switching to Manual — please browse.", "error")
                self.after(0, lambda: self.ref_mode.set("manual"))
                self.after(0, self._on_ref_mode_change)
                self.after(100, self._browse_reference)
                self._restore_reset_btn()
                self._set_btn_state(self.btn_scan, True)
            except Exception as e:
                self._scanning = False
                self.after(0, self._stop_thinking)
                self._log(f"✗ {e}", "error")
                self._restore_reset_btn()
                self._set_btn_state(self.btn_scan, True)
        threading.Thread(target=task, daemon=True).start()


    def _do_export(self):
        output = self.output_dir.get().strip()
        if not output:
            messagebox.showerror("Missing Info", "Please set an output folder.")
            return

        # Prepare list synchronously on main thread (fast, no DaVinci API calls)
        try:
            if not self.export_list:
                self._suppress_progress = True
                self.export_list = self.engine.prepare_export(
                    self.show_code.get().strip(),
                    self.show_acronym.get().strip(),
                    self.export_date.get().strip()
                )
                self._suppress_progress = False

            # Apply disabled episode filter
            disabled = self._disabled_episodes.copy()
            if disabled:
                self.export_list = [
                    item for item in self.export_list
                    if item.get("episode_code") not in disabled
                ]
                self._log(f"Skipping episodes: {', '.join(sorted(disabled))}.", "muted")
                self._log(f"  Update: {len(self.export_list)} plates to export.", "muted")

            # Sync to engine
            self.engine.export_list = self.export_list
            self._update_plate_count()

        except Exception as e:
            self._suppress_progress = False
            self._log(f"✗ Failed to prepare clips: {e}", "error")
            return

        # Now start the actual export
        self._run_export_task(output, start_index=0)

    def _browse_reference(self):
        path = filedialog.askopenfilename(
            title="Select Reference Video",
            filetypes=[("Video files", "*.mov *.mp4 *.mxf *.avi"), ("All files", "*.*")]
        )
        if path:
            self.reference_video_path.set(path)
            self._log(f"Ref video set to: {os.path.basename(path)}.", "success")

    def _browse_output(self):
        path = filedialog.askdirectory(title="Select Output Folder")
        if path:
            self.output_dir.set(path)

    def _log(self, message, tag=None):
        def _do():
            self.log_box.config(state="normal")
            self.log_box.insert("end", message + "\n", tag or "")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _do)

    def _update_progress(self, pct, msg=None):
        if getattr(self, '_suppress_progress', False):
            return
        import time
        now = time.time()
        # Throttle UI updates to max once per 150ms to keep main thread free
        last = getattr(self, '_last_progress_update', 0)
        if now - last < 0.15 and msg and msg.startswith("Exporting Plate:"):
            return  # Skip this update, UI already recent enough
        self._last_progress_update = now

        def _do():
            self.progress_var.set(pct)
            if msg and msg.startswith("Exporting Plate:"):
                if hasattr(self, 'progress_status'):
                    self.progress_status.config(text=msg)
            elif msg and not getattr(self, '_scanning', False):
                self._log(msg, "muted")
        self.after(0, _do)

    def _set_progress(self, pct):
        self.after(0, lambda: self.progress_var.set(pct))

    def _update_episode_list(self):
        def _do():
            for w in self._ep_tag_widgets:
                try: w.destroy()
                except: pass
            self._ep_tag_widgets.clear()

            # Build tags in wrapping rows
            tw, th, r, gap = 72, 24, 6, 5
            # Get available width - default to 600 if not yet rendered
            avail = max(self.ep_tags_frame.winfo_width() - 16, 400)
            cols = max(1, avail // (tw + gap))

            row_frame = None
            for idx, m in enumerate(self.episode_markers):
                if idx % cols == 0:
                    row_frame = tk.Frame(self.ep_tags_frame, bg="#252525")
                    row_frame.pack(fill="x", pady=(0, 2))
                    self._ep_tag_widgets.append(row_frame)

                ep = m.episode_code
                disabled = ep in self._disabled_episodes
                locked = getattr(self, '_export_started', False)
                if locked:
                    tag_fill = "#3d3d3d" if disabled else "#5a4a1a"
                    text_fill = "#888888"
                    toggle_char = "⊘"
                    text_fill = "#888888" if disabled else "#666666"
                else:
                    tag_fill = "#3d3d3d" if disabled else ACCENT
                    text_fill = "#888888" if disabled else "#000000"
                    toggle_char = "+" if disabled else "×"

                tag = tk.Canvas(row_frame, bg="#252525", highlightthickness=0,
                               width=tw, height=th)
                tag.pack(side="left", padx=(0, gap), pady=2)

                x1, y1, x2, y2 = 0, 0, tw, th
                pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r,
                       x2,y2-r, x2,y2, x2-r,y2, x1+r,y2,
                       x1,y2, x1,y2-r, x1,y1+r, x1,y1]
                tag.create_polygon(pts, fill=tag_fill, outline="", smooth=True)
                tag.create_text(12, th//2, text=toggle_char,
                               font=("SF Pro Display", 12, "bold"), fill=text_fill)
                tag.create_line(22, 4, 22, th-4, fill=text_fill)
                tag.create_text(22 + (tw-22)//2, th//2, text=ep,
                               font=("SF Pro Display", 10, "bold"), fill=text_fill)

                def _on_click(e, episode=ep):
                    if getattr(self, '_export_started', False):
                        return  # Locked during export
                    if episode in self._disabled_episodes:
                        self._disabled_episodes.discard(episode)
                    else:
                        self._disabled_episodes.add(episode)
                    self._update_episode_list()
                    self._update_plate_count()

                tag.bind("<ButtonRelease-1>", _on_click)
                self._ep_tag_widgets.append(tag)

        self.after(0, _do)

    def _toggle_all_episodes(self):
        """Toggle all episodes on or off."""
        if getattr(self, '_export_started', False):
            return
        if self._all_disabled:
            # Enable all
            self._disabled_episodes.clear()
            self._all_disabled = False
            self.btn_toggle_all.config(text="DISABLE ALL", bg=BG_INPUT, fg=TEXT_PRIMARY)
        else:
            # Disable all
            self._disabled_episodes = set(m.episode_code for m in self.episode_markers)
            self._all_disabled = True
            self.btn_toggle_all.config(text="ENABLE ALL", bg=BG_INPUT, fg=TEXT_PRIMARY)
        self._update_episode_list()
        self._update_plate_count()

    def _update_plate_count(self):
        if not self.export_list:
            return
        disabled = getattr(self, '_disabled_episodes', set())
        count = sum(1 for item in self.export_list
                   if item.get("episode_code", "") not in disabled)
        self.plates_count_label.config(text=str(count))

    def _set_btn_state(self, btn, enabled):
        def _do():
            btn._enabled = enabled
            if isinstance(btn, tk.Canvas):
                # Rounded canvas button
                if enabled:
                    bg = ACCENT if btn._accent else "#444444"
                    fg = "#000000" if btn._accent else "#FFFFFF"
                else:
                    bg = "#5a4a1a" if btn._accent else "#333333"
                    fg = "#888888" if btn._accent else "#555555"
                btn._bg = bg
                btn._fg = fg
                btn._draw(bg, fg)
                btn.config(cursor="" if enabled else "arrow")
                btn.unbind("<Enter>")
                btn.unbind("<Leave>")
                btn.unbind("<Button-1>")
                if enabled:
                    hover = ACCENT_HOVER if btn._accent else "#555555"
                    press = "#c07010" if btn._accent else "#666666"
                    btn.bind("<Enter>", lambda e, b=btn, h=hover, f=fg: b._draw(h, f))
                    btn.bind("<Leave>", lambda e, b=btn, f=fg: b._draw(b._bg, f))
                    btn.bind("<ButtonPress-1>", lambda e, b=btn, p=press, f=fg: b._draw(p, f))
                    def _on_release(e, b=btn, f=fg):
                        b._draw(b._bg, f)
                        b._command()
                    btn.bind("<ButtonRelease-1>", _on_release)
            else:
                # Legacy Label button
                if enabled:
                    bg = ACCENT if btn._accent else "#444444"
                    fg = "#000000" if btn._accent else "#FFFFFF"
                    btn._bg = bg
                    btn.config(bg=bg, fg=fg, cursor="")
                    def on_enter(e, b=btn): b.config(bg=ACCENT_HOVER if b._accent else "#555555")
                    def on_leave(e, b=btn): b.config(bg=b._bg)
                    def on_click(e, b=btn): b._command()
                    btn.bind("<Enter>", on_enter)
                    btn.bind("<Leave>", on_leave)
                    btn.bind("<Button-1>", on_click)
                else:
                    bg = "#5a4a1a" if btn._accent else "#333333"
                    fg = "#888888" if btn._accent else "#555555"
                    btn.config(bg=bg, fg=fg, cursor="arrow")
                    btn.unbind("<Button-1>")
                    btn.unbind("<Enter>")
                    btn.unbind("<Leave>")
        self.after(0, _do)


if __name__ == "__main__":
    import traceback
    try:
        app = VFXExporterApp()
        app.mainloop()
    except Exception as e:
        try:
            import tkinter.messagebox as mb
            mb.showerror("Startup Error", traceback.format_exc())
        except:
            traceback.print_exc()
